import hashlib
from io import BytesIO
from types import SimpleNamespace

from django.db import IntegrityError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.exceptions import APIException, PermissionDenied, ValidationError

from asn.models import AsnDetailModel, AsnListModel
from asn.views import (
    AsnArrivalConfirmView,
    AsnEtaUpdateView,
    AsnPreLoadViewSet,
    AsnPreSortViewSet,
    AsnSortedViewSet,
    MoveToBinViewSet,
)
from dn.models import DnDetailModel, DnListModel
from asn.serializers import ASNListGetSerializer
from binset.models import ListModel as Bin
from driver.models import ListModel as Driver
from stock.models import StockBinModel, StockListModel
from staging.models import StagingAssignment
from staff.models import ListModel as Staff
from supplier.models import ListModel as Supplier
from warehouse.models import ListModel as Warehouse
from goods.models import ListModel as Goods
from customer.models import ListModel as Customer
from userprofile.models import Users

from .models import (
    ACCEPT_FOR_PUTAWAY,
    HOLD_QUARANTINE,
    REPAIR_REWORK,
    REJECT_RETURN,
    AsnSerialRecord,
    AgentCommandPreview,
    EntityProvenance,
    MailboxSyncRun,
    MailboxSyncState,
    OperationAudit,
    PackListDocument,
    PackListImportBatch,
    PackListLine,
    SourceEvidence,
    SourceAttachment,
    SourceExtraction,
    SourceIntakeEvent,
    SourceIntakeRecord,
)
from .views import (
    SerialExceptionResolveView,
    SerialExceptionMoveView,
    AgentCommandPreviewView,
    PackListPreviewView,
    SerialImportPreviewView,
    _create_pack_list,
    _receiving_started,
    _scan,
    _serial_rows_from_workbook,
    _summary,
    WebWorkflowApproveView,
    WebWorkflowPreviewView,
    AgentCommandApproveView,
    MailboxSyncRunCompleteView,
    MailboxSyncRunCreateView,
    MailboxSyncStateView,
    SourceCaptureView,
    SourceIntakeDetailView,
    SourceIntakeListView,
    SourceIntakeUpdateView,
)
from .agent import (
    agent_roles_for_operation,
    approve_web_preview,
    complete_preview,
    consume_preview,
    consume_web_preview,
    create_source_capture,
    create_preview,
    create_web_preview,
    request_payload,
)
from .intake import ensure_source_intake_record, update_source_intake
from .permissions import AgentPreviewPermission


class AgentPreviewPermissionTests(TestCase):
    def request(self, staff_id=7, is_admin=False, operator='7', agent=True):
        return SimpleNamespace(
            user=SimpleNamespace(is_authenticated=True),
            auth=SimpleNamespace(openid='tenant', staff_id=staff_id, is_admin=is_admin),
            META={
                'HTTP_OPERATOR': operator,
                'HTTP_X_AGENT_CLIENT': 'greaterwms-cli' if agent else 'browser',
            },
        )

    def test_authenticated_outbound_operator_can_reach_agent_preview(self):
        self.assertTrue(AgentPreviewPermission().has_permission(self.request(), None))

    def test_agent_preview_rejects_operator_impersonation(self):
        self.assertFalse(AgentPreviewPermission().has_permission(self.request(operator='8'), None))

    def test_agent_preview_rejects_non_agent_requests(self):
        self.assertFalse(AgentPreviewPermission().has_permission(self.request(agent=False), None))

    def test_warehouse_can_preview_inbound_but_not_admin_workflows(self):
        self.assertIn('warehouse', agent_roles_for_operation('asn.create'))
        self.assertNotIn('warehouse', agent_roles_for_operation('tenant.cleanup'))
        self.assertNotIn('warehouse', agent_roles_for_operation('outbound.cancel_intransit'))

    def test_warehouse_can_preview_and_consume_inbound_command(self):
        operator = Staff.objects.create(
            staff_name='warehouse-operator',
            staff_type='Warehouse',
            openid='tenant',
        )
        payload = {
            'container_tracking': 'SIM-WAREHOUSE-001',
            'creater': operator.staff_name,
        }
        request = self.request(staff_id=operator.id, operator=str(operator.id))
        request.data = payload

        preview = create_preview(request, 'asn.create', payload)
        request.data = {
            'confirmation_token': preview['confirmation_token'],
            'idempotency_key': 'warehouse-asn-create-001',
        }
        command, replay = consume_preview(request, 'asn.create', payload)

        self.assertIsNotNone(command)
        self.assertIsNone(replay)
        self.assertEqual(command.operation, 'asn.create')
        self.assertEqual(command.created_by, str(operator.id))


class SourceProvenanceWorkflowTests(TestCase):
    def setUp(self):
        self.operator = Staff.objects.create(
            staff_name='web-operator',
            staff_type='Warehouse',
            openid='tenant',
        )
        Users.objects.create(
            name=self.operator.staff_name,
            openid='tenant',
            appid='test-app',
            t_code='test-tenant',
            ip='127.0.0.1',
        )

    def request(self, data=None, surface='web', client='browser'):
        return SimpleNamespace(
            user=SimpleNamespace(is_authenticated=True),
            auth=SimpleNamespace(
                openid='tenant',
                staff_id=self.operator.id,
                staff_name=self.operator.staff_name,
                staff_type=self.operator.staff_type,
                is_admin=False,
            ),
            META={
                'HTTP_OPERATOR': str(self.operator.id),
                'HTTP_X_AGENT_CLIENT': client,
                'HTTP_X_AGENT_SURFACE': surface,
            },
            data=data or {},
            GET={},
            query_params={},
            method='POST',
        )

    def test_email_capture_creates_source_intake_and_attachment_records(self):
        request = self.request(
            data={
                'source_type': 'EMAIL',
                'operation': 'external.instruction',
                'content_hash': 'a' * 64,
                'sync_run_id': '',
                'metadata': {
                    'mailbox_account': 'sales@example.com',
                    'message_id': '<message-1@example.com>',
                    'thread_id': 'thread-1',
                    'sender_name': 'Mark Tang',
                    'sender_email': 'mark@example.com',
                    'subject': 'Inbound Notice TRHU4217950',
                    'document_type': 'Inbound Notice',
                    'business_operation': 'inbound',
                    'external_reference': 'TRHU4217950',
                    'attachments': [{
                        'name': 'inbound-list.xlsx',
                        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        'sha256': 'b' * 64,
                        'size': 1024,
                        'security_status': 'STORED',
                    }],
                },
                'extracted_fields': [{
                    'field_name': 'container_no',
                    'raw_value': 'TRHU4217950',
                    'normalized_value': 'TRHU4217950',
                    'source_location': 'email body',
                    'confidence': 0.98,
                }],
            },
            surface='ai',
            client='greaterwms-ai',
        )
        response = SourceCaptureView().post(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(SourceEvidence.objects.count(), 1)
        self.assertEqual(SourceIntakeRecord.objects.count(), 1)
        self.assertEqual(SourceAttachment.objects.count(), 1)
        intake = SourceIntakeRecord.objects.get()
        self.assertEqual(intake.operation, SourceIntakeRecord.INBOUND)
        self.assertEqual(intake.document_type, SourceIntakeRecord.INBOUND_NOTICE)
        self.assertEqual(intake.external_reference, 'TRHU4217950')
        self.assertEqual(intake.status, SourceIntakeRecord.CAPTURED)
        self.assertEqual(SourceIntakeEvent.objects.filter(intake=intake).count(), 1)

    def test_duplicate_email_capture_is_recorded_without_duplicate_source(self):
        data = {
            'source_type': 'EMAIL',
            'operation': 'external.instruction',
            'content_hash': 'c' * 64,
            'metadata': {
                'mailbox_account': 'sales@example.com',
                'message_id': '<message-duplicate@example.com>',
                'document_type': 'Pack List',
                'business_operation': 'inbound',
            },
        }
        first_request = self.request(data=dict(data), surface='ai', client='greaterwms-ai')
        first = SourceCaptureView().post(first_request)
        second_request = self.request(data=dict(data), surface='ai', client='greaterwms-ai')
        second = SourceCaptureView().post(second_request)

        self.assertFalse(first.data['duplicate'])
        self.assertTrue(second.data['duplicate'])
        self.assertEqual(SourceEvidence.objects.count(), 1)
        self.assertEqual(SourceIntakeRecord.objects.count(), 1)
        self.assertEqual(SourceIntakeRecord.objects.get().status, SourceIntakeRecord.DUPLICATE)
        self.assertEqual(SourceIntakeEvent.objects.filter(event_type='DUPLICATE').count(), 1)

    def test_source_intake_status_transitions_are_logged_and_invalid_transition_is_rejected(self):
        source = SourceEvidence.objects.create(
            openid='tenant',
            source_type=SourceEvidence.EMAIL,
            operation='external.instruction',
            content_hash='d' * 64,
            mailbox_account='sales@example.com',
            message_id='<message-transition@example.com>',
        )
        intake, _ = ensure_source_intake_record(source)
        update_source_intake(intake, {'status': SourceIntakeRecord.ANALYZING}, actor_type='AI_AGENT')
        update_source_intake(intake, {
            'status': SourceIntakeRecord.READY_FOR_PREVIEW,
            'next_action': 'Create inbound preview',
        }, actor_type='AI_AGENT')

        with self.assertRaises(ValidationError):
            update_source_intake(intake, {'status': SourceIntakeRecord.COMPLETED}, actor_type='AI_AGENT')
        self.assertEqual(intake.events.count(), 3)

    def test_mailbox_sync_run_is_closed_with_counters(self):
        request = self.request(
            data={
                'mailbox_account': 'sales@example.com',
                'automation_run_id': 'codex-20260819-01',
            },
            surface='ai',
            client='greaterwms-ai',
        )
        created = MailboxSyncRunCreateView().post(request)
        run_id = created.data['id']
        complete_request = self.request(
            data={
                'status': 'PARTIAL',
                'fetched_count': 4,
                'captured_count': 2,
                'duplicate_count': 1,
                'review_count': 1,
                'failed_count': 0,
                'cursor_after': 'cursor-2',
            },
            surface='ai',
            client='greaterwms-ai',
        )
        completed = MailboxSyncRunCompleteView().post(complete_request, run_id)

        self.assertEqual(completed.data['status'], MailboxSyncRun.PARTIAL)
        run = MailboxSyncRun.objects.get(id=run_id)
        self.assertEqual(run.fetched_count, 4)
        self.assertEqual(run.cursor_after, 'cursor-2')

    def test_mailbox_sync_run_rejects_unknown_trigger_source(self):
        request = self.request(
            data={
                'mailbox_account': 'sales@example.com',
                'trigger_source': 'UNKNOWN_SCHEDULER',
            },
            surface='ai',
            client='greaterwms-ai',
        )
        with self.assertRaises(ValidationError):
            MailboxSyncRunCreateView().post(request)

    def test_mailbox_sync_cursor_advances_only_after_success(self):
        first_request = self.request(
            data={
                'mailbox_account': 'cursor@example.com',
                'automation_run_id': 'cursor-run-1',
            },
            surface='ai',
            client='greaterwms-ai',
        )
        first = MailboxSyncRunCreateView().post(first_request)
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.data['cursor_before'], '')

        partial = MailboxSyncRunCompleteView().post(
            self.request(
                data={
                    'status': 'PARTIAL',
                    'cursor_after': '100|message-1',
                    'failed_count': 1,
                },
                surface='ai',
                client='greaterwms-ai',
            ),
            first.data['id'],
        )
        self.assertEqual(partial.status_code, 200)
        state = MailboxSyncState.objects.get(openid='tenant', mailbox_account='cursor@example.com')
        self.assertEqual(state.cursor, '')

        second = MailboxSyncRunCreateView().post(
            self.request(
                data={
                    'mailbox_account': 'cursor@example.com',
                    'automation_run_id': 'cursor-run-2',
                },
                surface='ai',
                client='greaterwms-ai',
            )
        )
        self.assertEqual(second.status_code, 201)
        self.assertEqual(second.data['cursor_before'], '')
        completed = MailboxSyncRunCompleteView().post(
            self.request(
                data={
                    'status': 'SUCCEEDED',
                    'cursor_after': '200|message-2',
                },
                surface='ai',
                client='greaterwms-ai',
            ),
            second.data['id'],
        )
        self.assertEqual(completed.data['state_cursor'], '200|message-2')

        third = MailboxSyncRunCreateView().post(
            self.request(
                data={
                    'mailbox_account': 'cursor@example.com',
                    'automation_run_id': 'cursor-run-3',
                },
                surface='ai',
                client='greaterwms-ai',
            )
        )
        self.assertEqual(third.status_code, 201)
        self.assertEqual(third.data['cursor_before'], '200|message-2')

    def test_mailbox_sync_rejects_concurrent_run_and_exposes_state(self):
        first = MailboxSyncRunCreateView().post(
            self.request(
                data={'mailbox_account': 'locked@example.com'},
                surface='ai',
                client='greaterwms-ai',
            )
        )
        self.assertEqual(first.status_code, 201)

        blocked = MailboxSyncRunCreateView().post(
            self.request(
                data={'mailbox_account': 'locked@example.com'},
                surface='ai',
                client='greaterwms-ai',
            )
        )
        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(blocked.data['code'], 'MAILBOX_SYNC_IN_PROGRESS')

        state = MailboxSyncStateView().get(SimpleNamespace(
            auth=SimpleNamespace(openid='tenant', staff_id=self.operator.id, staff_type=self.operator.staff_type),
            META={'HTTP_OPERATOR': str(self.operator.id), 'HTTP_X_AGENT_CLIENT': 'greaterwms-ai'},
            query_params={'mailbox_account': 'locked@example.com'},
        ))
        self.assertTrue(state.data['active'])
        self.assertEqual(state.data['active_run_id'], first.data['id'])

    def test_web_preview_creates_source_and_requires_approval(self):
        payload = {
            'header': {'supplier': 'Delta', 'creater': self.operator.staff_name},
            'detail': {
                'supplier': 'Delta',
                'goods_code': ['702-S'],
                'goods_qty': [2],
                'creater': self.operator.staff_name,
            },
        }
        request = self.request()
        preview = create_web_preview(request, 'asn.create', payload, page='inbound.asn')
        command = AgentCommandPreview.objects.get(id=preview['preview_id'])
        self.assertEqual(command.execution_surface, 'WEB')
        self.assertEqual(command.source_evidence.source_type, SourceEvidence.WEB_FORM)
        intake = SourceIntakeRecord.objects.get(source=command.source_evidence)
        self.assertEqual(intake.operation, SourceIntakeRecord.INBOUND)
        self.assertEqual(intake.status, SourceIntakeRecord.APPROVAL_REQUIRED)
        self.assertEqual(intake.events.filter(event_type='WEB_PREVIEW_CREATED').count(), 1)
        self.assertEqual(command.status, AgentCommandPreview.PENDING)

        request.data = {'web_preview_id': command.id}
        with self.assertRaises(ValidationError) as error:
            consume_web_preview(request, 'asn.create', payload['header'], 'header')
        self.assertEqual(error.exception.detail['code'], 'WEB_APPROVAL_REQUIRED')

        approve_web_preview(request, command.id)
        request.META['HTTP_WEB_WORKFLOW_STAGE'] = 'header'
        command, replay = consume_web_preview(request, 'asn.create', payload['header'], 'header')
        self.assertIsNotNone(command)
        self.assertIsNone(replay)

        request.META['HTTP_WEB_WORKFLOW_STAGE'] = 'detail'
        changed = dict(payload['detail'], goods_qty=[3])
        with self.assertRaises(ValidationError) as error:
            consume_web_preview(request, 'asn.detail.create', changed, 'detail')
        self.assertEqual(error.exception.detail['code'], 'WEB_PAYLOAD_CHANGED')

    def test_ai_external_preview_requires_source_evidence(self):
        request = self.request(
            data={},
            surface='ai',
            client='greaterwms-ai',
        )
        with self.assertRaises(ValidationError) as error:
            create_preview(
                request,
                'asn.create',
                {'supplier': 'Delta', 'creater': self.operator.staff_name},
            )
        self.assertEqual(error.exception.detail['code'], 'SOURCE_EVIDENCE_REQUIRED')

    def test_web_preview_rejects_roles_outside_the_outbound_operation_matrix(self):
        payload = {
            'header': {'customer': 'Delta', 'creater': self.operator.staff_name},
            'detail': {
                'customer': 'Delta',
                'goods_code': ['702-S'],
                'goods_qty': [1],
                'creater': self.operator.staff_name,
            },
        }
        for staff_type in ('QC', 'StockControl', 'Inbound'):
            self.operator.staff_type = staff_type
            self.operator.save(update_fields=['staff_type'])
            request = self.request(
                data={'operation': 'outbound.create', 'payload': payload},
            )
            with self.assertRaises(PermissionDenied):
                WebWorkflowPreviewView().post(request)
        self.assertEqual(AgentCommandPreview.objects.count(), 0)
        self.assertEqual(SourceEvidence.objects.count(), 0)

    def test_web_approval_rechecks_role_after_preview_creation(self):
        payload = {
            'header': {'customer': 'Delta', 'creater': self.operator.staff_name},
            'detail': {
                'customer': 'Delta',
                'goods_code': ['702-S'],
                'goods_qty': [1],
                'creater': self.operator.staff_name,
            },
        }
        request = self.request(
            data={'operation': 'outbound.create', 'payload': payload},
        )
        preview = WebWorkflowPreviewView().post(request)
        command = AgentCommandPreview.objects.get(id=preview.data['preview_id'])

        self.operator.staff_type = 'QC'
        self.operator.save(update_fields=['staff_type'])
        request.auth.staff_type = 'QC'
        with self.assertRaises(PermissionDenied):
            approve_web_preview(request, command.id)

        command.refresh_from_db()
        self.assertEqual(command.status, AgentCommandPreview.PENDING)
        self.assertFalse(DnListModel.objects.filter(openid='tenant').exists())

    def test_ai_external_preview_uses_structured_approval_without_token(self):
        Supplier.objects.create(
            supplier_name='Delta', supplier_city='Plano', supplier_address='A',
            supplier_contact='A', supplier_manager='A', creater='tester', openid='tenant',
        )
        Warehouse.objects.create(
            warehouse_name='Peak', warehouse_city='Lewisville', warehouse_address='A',
            warehouse_contact='A', warehouse_manager='A', creater='tester', openid='tenant',
        )
        Goods.objects.create(
            goods_code='702-S', goods_desc='Cooling system', goods_supplier='Delta',
            goods_unit='EA', goods_class='Equipment', goods_brand='Delta', goods_color='N/A',
            goods_shape='Box', goods_specs='Standard', goods_origin='US', creater='tester',
            bar_code='702-S-BAR', openid='tenant',
        )
        request = self.request(surface='ai', client='greaterwms-ai')
        source = create_source_capture(request, {
            'source_type': 'EMAIL',
            'operation': 'asn.create',
            'metadata': {'message_id': '<ai-test@example.com>'},
        })
        payload = {
            'header': {'supplier': 'Delta', 'creater': self.operator.staff_name},
            'detail': {
                'supplier': 'Delta',
                'goods_code': ['702-S'],
                'goods_qty': [2],
                'creater': self.operator.staff_name,
            },
        }
        preview = create_preview(
            request,
            'asn.create',
            payload,
            source_evidence_id=source.id,
        )
        self.assertNotIn('confirmation_token', preview)
        command = AgentCommandPreview.objects.get(id=preview['preview_id'])
        self.assertEqual(command.execution_surface, 'AI')
        self.assertEqual(command.confirmation_token_hash, '')
        request.data = {}
        response = AgentCommandApproveView().post(request, command.id)
        self.assertEqual(response.status_code, 201)
        command.refresh_from_db()
        self.assertEqual(command.status, AgentCommandPreview.EXECUTED)
        self.assertEqual(
            SourceIntakeRecord.objects.get(source=source).status,
            SourceIntakeRecord.COMPLETED,
        )
        self.assertTrue(
            OperationAudit.objects.filter(
                preview=command,
                execution_surface='AI',
                status=OperationAudit.SUCCEEDED,
            ).exists()
        )
        self.assertTrue(
            EntityProvenance.objects.filter(
                source=source,
                entity_type='ASN',
                field_name='supplier',
                entity_ref__startswith='ASN',
            ).exists()
        )

    def test_web_asn_preview_approval_writes_parent_and_detail_atomically(self):
        Supplier.objects.create(
            supplier_name='Delta', supplier_city='Plano', supplier_address='A',
            supplier_contact='A', supplier_manager='A', creater='tester', openid='tenant',
        )
        Warehouse.objects.create(
            warehouse_name='Peak', warehouse_city='Lewisville', warehouse_address='A',
            warehouse_contact='A', warehouse_manager='A', creater='tester', openid='tenant',
        )
        Goods.objects.create(
            goods_code='702-S', goods_desc='Cooling system', goods_supplier='Delta',
            goods_unit='EA', goods_class='Equipment', goods_brand='Delta', goods_color='N/A',
            goods_shape='Box', goods_specs='Standard', goods_origin='US', creater='tester',
            bar_code='702-S-BAR', openid='tenant',
        )
        payload = {
            'header': {'supplier': 'Delta', 'creater': self.operator.staff_name},
            'detail': {
                'supplier': 'Delta', 'goods_code': ['702-S'], 'goods_qty': [2],
                'creater': self.operator.staff_name,
            },
        }
        request = self.request(data={'operation': 'asn.create', 'payload': payload})
        preview_response = WebWorkflowPreviewView().post(request)
        request.data = {}
        result = WebWorkflowApproveView().post(request, preview_response.data['preview_id'])
        self.assertEqual(result.status_code, 201)
        self.assertEqual(AsnListModel.objects.filter(openid='tenant', supplier='Delta').count(), 1)
        self.assertEqual(AsnDetailModel.objects.filter(openid='tenant', goods_code='702-S').count(), 1)
        self.assertEqual(SourceEvidence.objects.filter(openid='tenant', status=SourceEvidence.USED).count(), 1)
        self.assertEqual(
            SourceIntakeRecord.objects.get(source__source_type=SourceEvidence.WEB_FORM).status,
            SourceIntakeRecord.COMPLETED,
        )

    def test_web_outbound_preview_approval_writes_parent_and_detail(self):
        Customer.objects.create(
            customer_name='Delta', customer_city='Plano', customer_address='A',
            customer_contact='A', customer_manager='A', creater='tester', openid='tenant',
        )
        Warehouse.objects.create(
            warehouse_name='Peak', warehouse_city='Lewisville', warehouse_address='A',
            warehouse_contact='A', warehouse_manager='A', creater='tester', openid='tenant',
        )
        Goods.objects.create(
            goods_code='702-S', goods_desc='Cooling system', goods_supplier='Delta',
            goods_unit='EA', goods_class='Equipment', goods_brand='Delta', goods_color='N/A',
            goods_shape='Box', goods_specs='Standard', goods_origin='US', creater='tester',
            bar_code='702-S-BAR', openid='tenant', goods_price=100,
        )
        payload = {
            'header': {
                'customer': 'Delta', 'creater': self.operator.staff_name,
                'picking_mode': 'SKU_QTY',
            },
            'detail': {
                'customer': 'Delta', 'goods_code': ['702-S'], 'goods_qty': [1],
                'creater': self.operator.staff_name,
            },
        }
        request = self.request(data={'operation': 'outbound.create', 'payload': payload})
        preview_response = WebWorkflowPreviewView().post(request)
        request.data = {}
        result = WebWorkflowApproveView().post(request, preview_response.data['preview_id'])
        self.assertEqual(result.status_code, 201)
        self.assertEqual(DnListModel.objects.filter(openid='tenant', customer='Delta').count(), 1)
        self.assertEqual(DnDetailModel.objects.filter(openid='tenant', goods_code='702-S').count(), 1)

    def test_email_to_asn_to_putaway_completes_inbound_without_inventory_drift(self):
        """Exercise the source-backed AI intake and the physical ASN inbound path."""
        supplier_name = 'Delta Electronics (USA) Inc.'
        container_tracking = 'TRHU4217950'
        email_hash = hashlib.sha256(b'simulated inbound email and attachments').hexdigest()
        Supplier.objects.create(
            supplier_name=supplier_name,
            supplier_city='Plano',
            supplier_address='601 Data Dr',
            supplier_contact='Receiving',
            supplier_manager='Mark Tang',
            creater='tester',
            openid='tenant',
        )
        Warehouse.objects.create(
            warehouse_name='Peak Smart Lewisville',
            warehouse_city='Lewisville',
            warehouse_address='1991 Lakepointe Dr, Dock #24',
            warehouse_contact='Receiving',
            warehouse_manager='Warehouse Manager',
            creater='tester',
            openid='tenant',
        )
        Goods.objects.create(
            goods_code='702-S',
            goods_desc='CALLAN-MSFT 144KW Cooling System',
            goods_supplier=supplier_name,
            goods_unit='EA',
            goods_class='Equipment',
            goods_brand='Delta',
            goods_color='N/A',
            goods_shape='Box',
            goods_specs='Standard',
            goods_origin='US',
            goods_cost=100,
            goods_price=100,
            creater='tester',
            bar_code='702-S-BAR',
            openid='tenant',
        )
        Driver.objects.create(
            driver_name='Tom',
            license_plate='SIM-TRUCK-01',
            contact='N/A',
            creater='tester',
            openid='tenant',
        )
        Bin.objects.create(
            bin_name='A1-01',
            bin_size='STD',
            bin_property='Normal',
            location_role='STORAGE',
            staging_flow='',
            creater='tester',
            bar_code='SIM-BIN-A1-01',
            openid='tenant',
        )

        ai_request = self.request(surface='ai', client='greaterwms-ai')
        source_payload = {
            'source_type': 'EMAIL',
            'operation': 'asn.create',
            'mailbox_account': 'sales@texasranchenergy.com',
            'metadata': {
                'mailbox_account': 'sales@texasranchenergy.com',
                'sender': 'Mark Tang',
                'recipients': ['Peak Smart Logistics'],
                'cc': ['DEUS Receiving'],
                'subject': 'Delivery Request TRHU4217950, Delta IRHX Mesh & Fittings',
                'message_id': '<sim-inbound-001@texasranchenergy.com>',
                'thread_id': '<sim-thread-001@texasranchenergy.com>',
                'sent_at': '2026-08-18T12:24:00-05:00',
                'received_at': '2026-08-18T12:24:00-05:00',
                'folder': 'Inbox',
                'document_classification': 'Inbound List / Pack List reference',
                'attachments': [{
                    'name': 'inbound-list.pdf',
                    'mime_type': 'application/pdf',
                    'size': 18432,
                    'sha256': hashlib.sha256(b'inbound-list.pdf').hexdigest(),
                    'source_location': 'attachment, page 1',
                }],
                'conflicts': {
                    'delivery_address': [
                        'Delta Electronics, 601 Data Dr, Plano, TX',
                        'Peak Smart Logistics, 1991 Lakepointe Dr, Dock #24, Lewisville, TX',
                    ],
                },
            },
            'content_hash': email_hash,
            'storage_uri': 'test://encrypted-object-store/inbound-email-001.eml',
            'storage_size': 18432,
            'ai_session_id': 'sim-ai-session-inbound-001',
            'extracted_fields': [
                {
                    'field_name': 'container_tracking',
                    'raw_value': 'TRHU4217950',
                    'normalized_value': container_tracking,
                    'source_location': 'email body, Delivery Request subject',
                    'confidence': '0.9900',
                    'human_confirmed': True,
                    'used_for_write': True,
                },
                {
                    'field_name': 'supplier',
                    'raw_value': 'Delta Electronics (USA) Inc.',
                    'normalized_value': supplier_name,
                    'source_location': 'attachment, page 1, Client',
                    'confidence': '0.9900',
                    'human_confirmed': True,
                    'used_for_write': True,
                },
                {
                    'field_name': 'expected_arrival_at',
                    'raw_value': '2026-08-16',
                    'normalized_value': '2026-08-16T00:00:00-05:00',
                    'source_location': 'attachment, page 1, ETA',
                    'confidence': '0.9800',
                    'human_confirmed': True,
                    'used_for_write': True,
                },
                {
                    'field_name': 'goods_code',
                    'raw_value': 'CL SAC144AD702-S',
                    'normalized_value': '702-S',
                    'source_location': 'attachment, page 1, SKU',
                    'confidence': '0.9700',
                    'human_confirmed': True,
                    'used_for_write': True,
                },
                {
                    'field_name': 'goods_qty',
                    'raw_value': '2',
                    'normalized_value': '2',
                    'source_location': 'attachment, page 1, Item Qty',
                    'confidence': '0.9900',
                    'human_confirmed': True,
                    'used_for_write': True,
                },
            ],
        }
        source = create_source_capture(ai_request, source_payload)
        duplicate_source = create_source_capture(ai_request, source_payload)
        self.assertEqual(duplicate_source.id, source.id)
        self.assertEqual(SourceEvidence.objects.filter(openid='tenant').count(), 1)
        self.assertEqual(SourceExtraction.objects.filter(source=source).count(), 5)

        payload = {
            'header': {
                'supplier': supplier_name,
                'container_tracking': container_tracking,
                'creater': self.operator.staff_name,
            },
            'detail': {
                'supplier': supplier_name,
                'goods_code': ['702-S'],
                'goods_qty': [2],
                'creater': self.operator.staff_name,
            },
        }
        preview = create_preview(
            ai_request,
            'asn.create',
            payload,
            source_evidence_id=source.id,
        )
        self.assertNotIn('confirmation_token', preview)
        result = AgentCommandApproveView().post(ai_request, preview['preview_id'])
        self.assertEqual(result.status_code, 201)
        asn = AsnListModel.objects.get(openid='tenant', container_tracking=container_tracking)
        asn_detail = AsnDetailModel.objects.get(openid='tenant', asn_code=asn.asn_code, goods_code='702-S')
        stock = StockListModel.objects.get(openid='tenant', goods_code='702-S')
        self.assertEqual(asn.asn_status, 1)
        self.assertEqual(asn_detail.goods_qty, 2)
        self.assertEqual(stock.goods_qty, 2)
        self.assertEqual(stock.asn_stock, 2)

        def cli_request(data):
            return self.request(data=data, surface='cli', client='greaterwms-cli')

        def cli_preview(operation, data, resource_id):
            request = cli_request(data)
            return request, create_preview(
                request,
                operation,
                data,
                resource_id=str(resource_id),
                asn_code=asn.asn_code,
            )

        eta_data = {
            'expected_arrival_at': '2026-08-16T00:00:00',
            'source': 'CUSTOMER_EMAIL',
            'note': 'ETA extracted from inbound list attachment',
        }
        eta_request, eta_preview = cli_preview('asn.eta', eta_data, asn.id)
        eta_request.data = dict(
            eta_data,
            confirmation_token=eta_preview['confirmation_token'],
            idempotency_key='sim-inbound-eta-001',
        )
        eta_response = AsnEtaUpdateView().post(eta_request, asn.id)
        self.assertEqual(eta_response.status_code, 200)

        arrival_data = {
            'actual_arrival_at': '2026-08-18T13:30:00',
            'source': 'WAREHOUSE',
            'note': 'Truck arrived; unload authorized',
        }
        arrival_request, arrival_preview = cli_preview('asn.arrival', arrival_data, asn.id)
        arrival_request.data = dict(
            arrival_data,
            confirmation_token=arrival_preview['confirmation_token'],
            idempotency_key='sim-inbound-arrival-001',
        )
        arrival_response = AsnArrivalConfirmView().post(arrival_request, asn.id)
        self.assertEqual(arrival_response.status_code, 200)

        unload_data = {
            'unload_driver': 'Tom',
            'staging_bins': ['STAGE-LEFT-01', 'STAGE-LEFT-02'],
        }
        unload_request, unload_preview = cli_preview('asn.unload_start', unload_data, asn.id)
        unload_request.data = dict(
            unload_data,
            confirmation_token=unload_preview['confirmation_token'],
            idempotency_key='sim-inbound-unload-001',
        )
        unload_view = AsnPreLoadViewSet()
        unload_view.request = unload_request
        unload_view.action = 'create'
        unload_view.format_kwarg = None
        unload_response = unload_view.create(unload_request, asn.id)
        self.assertEqual(unload_response.status_code, 200)
        asn.refresh_from_db()
        self.assertEqual(asn.asn_status, 2)
        self.assertEqual(
            StagingAssignment.objects.filter(
                openid='tenant', reference_code=asn.asn_code,
                flow=StagingAssignment.INBOUND,
                status=StagingAssignment.RESERVED,
            ).count(),
            2,
        )

        finish_request, finish_preview = cli_preview('asn.unload_finish', {}, asn.id)
        finish_request.data = {
            'confirmation_token': finish_preview['confirmation_token'],
            'idempotency_key': 'sim-inbound-unload-finish-001',
        }
        finish_view = AsnPreSortViewSet()
        finish_view.request = finish_request
        finish_view.action = 'create'
        finish_view.format_kwarg = None
        finish_view.get_object = lambda: AsnListModel.objects.get(id=asn.id)
        finish_response = finish_view.create(finish_request, asn.id)
        self.assertEqual(finish_response.status_code, 200)
        asn.refresh_from_db()
        self.assertEqual(asn.asn_status, 3)
        self.assertEqual(
            StagingAssignment.objects.filter(
                openid='tenant', reference_code=asn.asn_code,
                flow=StagingAssignment.INBOUND,
                status=StagingAssignment.ACTIVE,
            ).count(),
            2,
        )

        receive_data = {
            'asn_code': asn.asn_code,
            'supplier': supplier_name,
            'goodsData': [{'goods_code': '702-S', 'goods_actual_qty': 2}],
        }
        receive_request, receive_preview = cli_preview('asn.receive', receive_data, asn.id)
        receive_request.data = dict(
            receive_data,
            confirmation_token=receive_preview['confirmation_token'],
            idempotency_key='sim-inbound-receive-001',
        )
        receive_view = AsnSortedViewSet()
        receive_view.request = receive_request
        receive_view.action = 'create'
        receive_view.format_kwarg = None
        receive_view.get_object = lambda: AsnListModel.objects.get(id=asn.id)
        receive_response = receive_view.create(receive_request, asn.id)
        self.assertEqual(receive_response.status_code, 200)
        asn.refresh_from_db()
        asn_detail.refresh_from_db()
        stock.refresh_from_db()
        self.assertEqual(asn.asn_status, 4)
        self.assertEqual(asn_detail.goods_actual_qty, 2)
        self.assertEqual(asn_detail.sorted_qty, 0)
        self.assertEqual(stock.pre_sort_stock, 0)
        self.assertEqual(stock.sorted_stock, 2)
        self.assertEqual(stock.goods_qty, 2)

        putaway_data = {
            'asn_code': asn.asn_code,
            'goods_code': '702-S',
            'qty': 2,
            'bin_name': 'A1-01',
            'putaway_driver': 'Tom',
        }
        putaway_request, putaway_preview = cli_preview('asn.putaway', putaway_data, asn_detail.id)
        putaway_request.data = dict(
            putaway_data,
            confirmation_token=putaway_preview['confirmation_token'],
            idempotency_key='sim-inbound-putaway-001',
        )
        putaway_view = MoveToBinViewSet()
        putaway_view.request = putaway_request
        putaway_view.action = 'create'
        putaway_view.format_kwarg = None
        putaway_view.get_object = lambda: AsnDetailModel.objects.get(id=asn_detail.id)
        putaway_response = putaway_view.create(putaway_request, asn_detail.id)
        self.assertEqual(putaway_response.status_code, 200)

        asn.refresh_from_db()
        asn_detail.refresh_from_db()
        stock.refresh_from_db()
        source.refresh_from_db()
        self.assertEqual(asn.asn_status, 5)
        self.assertEqual(asn_detail.asn_status, 5)
        self.assertEqual(asn_detail.sorted_qty, 2)
        self.assertEqual(asn.putaway_driver, 'Tom')
        self.assertEqual(stock.goods_qty, 2)
        self.assertEqual(stock.onhand_stock, 2)
        self.assertEqual(stock.can_order_stock, 2)
        self.assertEqual(stock.asn_stock, 0)
        self.assertEqual(
            StagingAssignment.objects.filter(
                openid='tenant', reference_code=asn.asn_code,
                flow=StagingAssignment.INBOUND,
                status=StagingAssignment.RELEASED,
            ).count(),
            2,
        )
        self.assertEqual(source.status, SourceEvidence.USED)
        self.assertTrue(
            EntityProvenance.objects.filter(
                source=source,
                entity_type='ASN',
                entity_ref=asn.asn_code,
                field_name='container_tracking',
            ).exists()
        )
        self.assertTrue(
            OperationAudit.objects.filter(
                source_evidence=source,
                operation='asn.create',
                status=OperationAudit.SUCCEEDED,
                execution_surface='AI',
            ).exists()
        )


class PackListWorkflowTests(TestCase):
    def setUp(self):
        self.openid = 'test-tenant'
        self.asn_code = 'ASN-TEST-01'
        AsnListModel.objects.create(
            asn_code=self.asn_code,
            asn_status=1,
            supplier='Test Customer',
            creater='tester',
            bar_code='BAR-TEST-01',
            openid=self.openid,
            transportation_fee={},
        )
        AsnDetailModel.objects.create(
            asn_code=self.asn_code,
            asn_status=1,
            supplier='Test Customer',
            goods_code='702-S',
            goods_desc='Test SKU',
            goods_qty=2,
            creater='tester',
            openid=self.openid,
        )

    def request(self):
        return SimpleNamespace(
            auth=SimpleNamespace(openid=self.openid),
            META={},
        )

    def agent_request(self, data=None, operator_id=None):
        if operator_id is None:
            operator_id = Staff.objects.create(
                openid=self.openid,
                staff_name='Inbound Operator',
                staff_type='Inbound',
            ).id
        return SimpleNamespace(
            auth=SimpleNamespace(openid=self.openid),
            META={
                'HTTP_X_AGENT_CLIENT': 'greaterwms-cli',
                'HTTP_OPERATOR': str(operator_id),
            },
            data=data or {},
        )

    def rows(self):
        return [{
            'goods_code': '702-S',
            'customer_goods_code': 'CUSTOMER-702',
            'customer_ssku': 'S-702',
            'package_type': 'PKG-01',
            'serial_number': '',
            'goods_qty': 2,
            'total_qty': 2,
            'goods_desc': 'Test SKU',
            'goods_weight': 10,
            'goods_volume': 1,
            'source_row': 2,
        }]

    def workbook_upload(self, headers, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        payload = BytesIO()
        workbook.save(payload)
        return SimpleUploadedFile(
            'inbound-smoke.xlsx',
            payload.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_pack_list_import_is_one_current_record_and_idempotent(self):
        document, batch, created = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='a' * 64,
            package_qty=2,
        )
        self.assertTrue(created)
        self.assertIsNotNone(batch)
        same, no_batch, was_created = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='a' * 64,
            package_qty=2,
        )
        self.assertEqual(document.id, same.id)
        self.assertIsNone(no_batch)
        self.assertFalse(was_created)
        self.assertEqual(PackListDocument.objects.filter(is_current=True).count(), 1)
        self.assertEqual(PackListLine.objects.filter(pack_list=document, is_current=True).count(), 1)

        with self.assertRaises(APIException) as error:
            _create_pack_list(
                self.openid,
                self.request(),
                self.asn_code,
                self.rows(),
                content_hash='b' * 64,
                package_qty=2,
            )
        self.assertEqual(error.exception.detail['code'], 'PACK_LIST_REPLACE_REQUIRED')

    def test_receiving_status_blocks_pack_list_replacement_before_quantity_is_entered(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        asn.asn_status = 3
        asn.save(update_fields=['asn_status'])

        self.assertTrue(_receiving_started(self.openid, self.asn_code))

    def test_missing_expected_serial_cannot_be_accepted_for_putaway(self):
        record = AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-MISSING',
            is_expected=True,
            is_received=False,
            status=AsnSerialRecord.EXPECTED,
        )
        request = self.request()
        request.data = {
            'id': record.id,
            'action': ACCEPT_FOR_PUTAWAY,
            'note': 'Incorrectly attempted to bypass missing SN',
        }

        with self.assertRaises(APIException) as error:
            SerialExceptionResolveView().post(request)

        self.assertEqual(error.exception.detail['code'], 'MISSING_SN_NOT_PUTAWAY_ELIGIBLE')

    def test_resolved_received_serial_exception_moves_stock_and_releases_staging(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        asn.asn_status = 4
        asn.save(update_fields=['asn_status'])
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.asn_status = 4
        detail.goods_actual_qty = 1
        detail.sorted_qty = 0
        detail.save(update_fields=['asn_status', 'goods_actual_qty', 'sorted_qty'])
        StockListModel.objects.create(
            openid=self.openid,
            goods_code='702-S',
            goods_desc='Test SKU',
            goods_qty=1,
            sorted_stock=1,
        )
        Bin.objects.create(
            openid=self.openid,
            bin_name='QC-HOLD-01',
            bin_size='STD',
            bin_property='Holding',
            location_role='STORAGE',
            staging_flow='NONE',
            creater='tester',
            bar_code='QC-HOLD-01',
        )
        record = AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-HOLD-001',
            is_expected=True,
            is_received=True,
            status=AsnSerialRecord.DAMAGED,
            exception_resolved=True,
            exception_resolution_action=HOLD_QUARANTINE,
            exception_resolution_location='QC-HOLD-01',
        )
        request = self.request()
        request.data = {'id': record.id, 'bin_name': 'QC-HOLD-01'}

        response = SerialExceptionMoveView().post(request)

        self.assertEqual(response.data['destination_bin'], 'QC-HOLD-01')
        record.refresh_from_db()
        detail.refresh_from_db()
        asn.refresh_from_db()
        stock = StockListModel.objects.get(openid=self.openid, goods_code='702-S')
        self.assertTrue(record.exception_moved)
        self.assertEqual(detail.sorted_qty, 1)
        self.assertEqual(detail.asn_status, 5)
        self.assertEqual(asn.asn_status, 5)
        self.assertEqual(stock.sorted_stock, 0)
        self.assertEqual(stock.onhand_stock, 1)
        self.assertEqual(StockBinModel.objects.get(goods_code='702-S').goods_qty, 1)

    def test_pack_list_defaults_to_ai_agent_source_and_audits_batch_source(self):
        document, batch, created = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='f' * 64,
        )

        self.assertTrue(created)
        self.assertEqual(document.source_type, 'AI_AGENT')
        self.assertEqual(batch.source_type, 'AI_AGENT')

    def test_agent_preview_token_is_payload_bound_and_idempotent(self):
        request = self.agent_request()
        preview = create_preview(
            request,
            'packlist.confirm',
            {'id': 123},
            resource_id='123',
        )
        execute_request = self.agent_request({
            'id': 123,
            'confirmation_token': preview['confirmation_token'],
            'idempotency_key': 'packlist-confirm-123-1',
        }, operator_id=request.META['HTTP_OPERATOR'])
        command, replay = consume_preview(
            execute_request,
            'packlist.confirm',
            request_payload(execute_request),
            resource_id='123',
        )
        self.assertIsNone(replay)
        complete_preview(command, {'detail': 'success'})

        replay_command, replay = consume_preview(
            execute_request,
            'packlist.confirm',
            request_payload(execute_request),
            resource_id='123',
        )
        self.assertEqual(replay, {'detail': 'success'})
        self.assertEqual(replay_command.id, command.id)

    def test_outbound_role_can_confirm_outbound_preview(self):
        operator = Staff.objects.create(
            openid=self.openid,
            staff_name='Outbound Operator',
            staff_type='Outbound',
        )
        payload = {
            'customer': 'Test Customer',
            'creater': 'Outbound Operator',
        }
        preview_request = self.agent_request(operator_id=operator.id)
        preview = create_preview(preview_request, 'outbound.create', payload)
        execute_request = self.agent_request({
            **payload,
            'confirmation_token': preview['confirmation_token'],
            'idempotency_key': 'outbound-create-test-1',
        }, operator_id=operator.id)

        command, replay = consume_preview(
            execute_request,
            'outbound.create',
            request_payload(execute_request),
        )
        self.assertIsNone(replay)
        complete_preview(command, {'detail': 'success'})

    def test_invalid_agent_token_is_a_client_error(self):
        request = self.agent_request({
            'id': 123,
            'confirmation_token': 'invalid-token',
            'idempotency_key': 'packlist-confirm-invalid-1',
        })

        with self.assertRaises(Exception) as raised:
            consume_preview(
                request,
                'packlist.confirm',
                request_payload(request),
                resource_id='123',
            )

        self.assertEqual(raised.exception.status_code, 400)

    def test_putaway_preview_reuses_final_putaway_gates(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        asn.asn_status = 4
        asn.save(update_fields=['asn_status'])
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.asn_status = 4
        detail.goods_actual_qty = 2
        detail.sorted_qty = 2
        detail.save(update_fields=['asn_status', 'goods_actual_qty', 'sorted_qty'])
        Driver.objects.create(
            openid=self.openid,
            driver_name='Tom',
            license_plate='TEST-001',
            contact='555-0001',
            creater='tester',
        )
        Bin.objects.create(
            openid=self.openid,
            bin_name='A1-01',
            bin_size='STD',
            bin_property='Normal',
            location_role='STORAGE',
            staging_flow='NONE',
            creater='tester',
            bar_code='A1-01',
        )
        StockListModel.objects.create(
            openid=self.openid,
            goods_code='702-S',
            goods_desc='Test SKU',
            goods_qty=2,
            sorted_stock=0,
        )
        request = self.agent_request({
            'operation': 'asn.putaway',
            'resource_id': str(detail.id),
            'asn_code': self.asn_code,
            'payload': {
                'asn_code': self.asn_code,
                'goods_code': '702-S',
                'qty': 1,
                'bin_name': 'A1-01',
                'putaway_driver': 'Tom',
            },
        })

        with self.assertRaises(Exception) as raised:
            AgentCommandPreviewView().post(request)

        self.assertEqual(raised.exception.status_code, 400)
        self.assertIn('remaining received quantity', str(raised.exception.detail))

    def test_pack_list_preview_returns_client_error_for_unknown_sku(self):
        request = self.request()
        request.data = {'asn_code': self.asn_code}
        request.FILES = {
            'file': self.workbook_upload(['SKU', 'Item Qty'], [['UNKNOWN-SKU', 1]]),
        }

        with self.assertRaises(Exception) as raised:
            PackListPreviewView().post(request)

        self.assertEqual(raised.exception.status_code, 400)
        self.assertIn('Goods Code is not part of this ASN', str(raised.exception.detail))

    def test_inspection_preview_returns_client_error_for_unscoped_scan_sheet(self):
        request = self.request()
        request.data = {'asn_code': self.asn_code}
        request.FILES = {
            'file': self.workbook_upload(['SKU#', 'SN#'], [['702-S', 'SN-001']]),
        }

        with self.assertRaises(Exception) as raised:
            SerialImportPreviewView().post(request, inspection=True)

        self.assertEqual(raised.exception.status_code, 400)
        self.assertIn('inbound_po or shipout_ref', str(raised.exception.detail))

    def test_summary_exposes_pending_pack_list_reconciliation(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='a' * 64,
            package_qty=2,
        )

        summary = _summary(self.openid, self.asn_code)
        row = summary['reconciliation_rows'][0]

        self.assertEqual(summary['reconciliation_status'], 'REVIEW')
        self.assertEqual(summary['customer_sn_status'], 'NOT_PROVIDED')
        self.assertEqual(row['goods_code'], '702-S')
        self.assertEqual(row['customer_goods_code'], 'CUSTOMER-702')
        self.assertEqual(row['pack_list_qty'], 2)
        self.assertEqual(row['received_qty'], 2)
        self.assertEqual(row['accepted_qty'], 2)
        self.assertEqual(row['variance'], 0)
        self.assertEqual(row['open_exception_count'], 0)
        self.assertEqual(row['result'], 'REVIEW')
        self.assertEqual(summary['receiving_summary']['status'], 'PASSED')

    def test_acceptance_workbook_reads_matching_rows_across_sheets_and_sections(self):
        workbook = Workbook()
        first = workbook.active
        first.title = 'Scan'
        first.append(['Inbound PO#', 'SKU#', 'SN#'])
        first.append(['PO-001', '702-S', 'SN-001'])
        first.append(['PO-002', '702-S', 'SN-002'])
        first.append(['SKU#', 'SN#', 'Result'])
        first.append(['702-S', 'SN-003', 'PASS'])
        second = workbook.create_sheet('Verification')
        second.append(['SKU', 'Serial Number', 'Status'])
        second.append(['702-S', 'SN-004', 'PASS'])
        payload = BytesIO()
        workbook.save(payload)

        rows = _serial_rows_from_workbook(payload.getvalue(), inbound_po='PO-001')

        self.assertEqual([(row['sheet'], row['row_number']) for row in rows], [('Scan', 2)])

    def test_acceptance_workbook_returns_no_rows_for_nonmatching_filter(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Inbound PO#', 'SKU#', 'SN#'])
        sheet.append(['PO-001', '702-S', 'SN-001'])
        payload = BytesIO()
        workbook.save(payload)

        rows = _serial_rows_from_workbook(payload.getvalue(), inbound_po='PO-MISSING')

        self.assertEqual(rows, [])

    def test_imported_qc_batch_cannot_report_passed_or_ready(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        PackListImportBatch.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
            status=PackListImportBatch.IMPORTED,
            row_count=2,
            matched_count=2,
        )

        summary = _summary(self.openid, self.asn_code)

        self.assertEqual(summary['qc_status'], 'PARTIAL')
        self.assertEqual(summary['receiving_summary']['status'], 'REVIEW')
        self.assertTrue(summary['qc_import_incomplete'])
        self.assertFalse(summary['qc_complete'])
        self.assertFalse(summary['ready_for_putaway'])

    def test_asn_serializer_handles_missing_pack_list(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)

        data = ASNListGetSerializer(asn, context={}).data

        self.assertEqual(data['pack_list_status'], 'NOT_RECEIVED')
        self.assertEqual(data['serial_acceptance']['status'], 'NOT_IMPORTED')

    def test_quantity_only_receipt_without_pack_list_is_ready_for_putaway(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        asn.actual_arrival_at = timezone.now()
        asn.asn_status = 3
        asn.save(update_fields=['actual_arrival_at', 'asn_status'])
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])

        receiving_review = ASNListGetSerializer(asn, context={}).data
        self.assertEqual(receiving_review['operational_status'], 'RECEIVING_REVIEW')
        self.assertEqual(receiving_review['next_action_code'], 'REVIEW_RECEIVING')

        asn.asn_status = 4
        asn.save(update_fields=['asn_status'])

        data = ASNListGetSerializer(asn, context={}).data
        summary = _summary(self.openid, self.asn_code)

        self.assertEqual(data['serial_acceptance']['status'], 'NOT_IMPORTED')
        self.assertTrue(data['serial_acceptance']['qc_complete'])
        self.assertEqual(data['operational_status'], 'READY_FOR_PUTAWAY')
        self.assertEqual(data['next_action_code'], 'ASSIGN_DRIVER_PUTAWAY')
        self.assertTrue(summary['qc_complete'])
        self.assertTrue(summary['ready_for_putaway'])

    def test_asn_serializer_exposes_operational_status_for_work_queue(self):
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)

        data = ASNListGetSerializer(asn, context={}).data
        self.assertEqual(data['operational_status'], 'PENDING_ARRIVAL')
        self.assertEqual(data['next_action_code'], 'SET_ETA')

        asn.actual_arrival_at = timezone.now()
        asn.asn_status = 4
        asn.save(update_fields=['actual_arrival_at', 'asn_status'])
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        for serial_number in ('SN-STATUS-001', 'SN-STATUS-002'):
            AsnSerialRecord.objects.create(
                openid=self.openid,
                asn_code=self.asn_code,
                goods_code='702-S',
                serial_number=serial_number,
                status=AsnSerialRecord.ACCEPTED,
                is_expected=True,
                is_received=True,
            )
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='status-pending',
            package_qty=2,
        )

        data = ASNListGetSerializer(asn, context={}).data
        self.assertEqual(data['operational_status'], 'PACK_LIST_REVIEW')
        self.assertEqual(data['next_action_code'], 'REVIEW_PACK_LIST')
        self.assertEqual(data['putaway_qty'], 0)

        damaged = AsnSerialRecord.objects.get(serial_number='SN-STATUS-002')
        damaged.status = AsnSerialRecord.DAMAGED
        damaged.damaged = True
        damaged.note = 'Outer packaging damage'
        damaged.save(update_fields=['status', 'damaged', 'note'])
        data = ASNListGetSerializer(asn, context={}).data
        self.assertEqual(data['operational_status'], 'QC_REVIEW_REQUIRED')
        self.assertEqual(data['next_action_code'], 'REVIEW_QC')

    def test_extra_scan_record_does_not_increase_received_or_putaway_qty(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        for serial_number in ('SN-702-001', 'SN-702-002'):
            AsnSerialRecord.objects.create(
                openid=self.openid,
                asn_code=self.asn_code,
                goods_code='702-S',
                serial_number=serial_number,
                status=AsnSerialRecord.ACCEPTED,
                is_expected=True,
                is_received=True,
            )
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-702-EXTRA',
            status=AsnSerialRecord.UNEXPECTED,
            is_expected=False,
            is_received=True,
            exception_resolved=True,
        )

        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        acceptance = ASNListGetSerializer(asn, context={}).data['serial_acceptance']
        summary = _summary(self.openid, self.asn_code)

        self.assertEqual(acceptance['actual_received_qty'], 2)
        self.assertEqual(acceptance['scan_record_count'], 3)
        self.assertEqual(acceptance['extra_scan_count'], 1)
        self.assertEqual(acceptance['accepted_for_putaway'], 2)
        self.assertEqual(acceptance['putaway_qty'], 2)
        self.assertEqual(summary['receiving_summary']['received_qty'], 2)
        self.assertEqual(summary['receiving_summary']['scan_record_count'], 3)
        self.assertEqual(summary['receiving_summary']['extra_scan_records'], 1)
        self.assertEqual(summary['receiving_summary']['putaway_qty'], 0)
        self.assertEqual(summary['total_accepted_for_putaway'], 2)
        self.assertEqual(summary['total_putaway_qty'], 0)
        self.assertEqual(summary['reconciliation_rows'][0]['received_qty'], 2)

    def test_summary_marks_reconciliation_exception_for_quantity_variance(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 1
        detail.save(update_fields=['goods_actual_qty'])
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='a' * 64,
            package_qty=2,
        )

        summary = _summary(self.openid, self.asn_code)

        self.assertEqual(summary['reconciliation_status'], 'EXCEPTION')
        self.assertEqual(summary['reconciliation_rows'][0]['variance'], -1)
        self.assertEqual(summary['reconciliation_rows'][0]['result'], 'EXCEPTION')

    def test_archived_pack_list_does_not_compete_with_current_record(self):
        current = PackListDocument.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            content_hash='c' * 64,
            is_current=True,
        )
        archived = PackListDocument.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            content_hash='d' * 64,
            is_current=False,
            status=PackListDocument.ARCHIVED,
        )
        self.assertEqual(current.asn_code, archived.asn_code)
        with self.assertRaises(IntegrityError):
            PackListDocument.objects.create(
                openid=self.openid,
                asn_code=self.asn_code,
                content_hash='e' * 64,
                is_current=True,
            )

    def test_explicit_replace_reuses_current_document_and_archives_old_lines(self):
        document, _, _ = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='a' * 64,
            package_qty=2,
        )
        replacement_rows = self.rows()
        replacement_rows[0]['customer_goods_code'] = 'CUSTOMER-702-REV2'
        replaced, batch, created = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            replacement_rows,
            content_hash='b' * 64,
            package_qty=2,
            replace=True,
        )
        self.assertFalse(created)
        self.assertIsNotNone(batch)
        self.assertEqual(document.id, replaced.id)
        self.assertEqual(replaced.version, 2)
        self.assertEqual(PackListDocument.objects.filter(is_current=True).count(), 1)
        self.assertEqual(PackListLine.objects.filter(pack_list=document, is_current=True).count(), 1)
        self.assertEqual(PackListLine.objects.filter(pack_list=document, is_current=False).count(), 1)

    def test_late_pack_list_is_a_new_reference_revision_after_receiving_started(self):
        rows = self.rows()
        rows[0]['serial_number'] = 'SN-702-001'
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            rows,
            content_hash='a' * 64,
            package_qty=2,
        )
        _scan(self.openid, self.request(), self.asn_code, '702-S', 'SN-702-001')
        self.assertTrue(AsnSerialRecord.objects.get(serial_number='SN-702-001').is_received)
        with self.assertRaises(APIException) as error:
            _create_pack_list(
                self.openid,
                self.request(),
                self.asn_code,
                self.rows(),
                content_hash='b' * 64,
                package_qty=2,
                replace=True,
            )
        self.assertEqual(error.exception.detail['code'], 'PACK_LIST_LATE_REFERENCE_REQUIRED')
        late, _, created = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            self.rows(),
            content_hash='b' * 64,
            package_qty=2,
            replace=True,
            late_reference=True,
        )
        self.assertTrue(created)
        self.assertTrue(late.late_reference)
        self.assertTrue(late.is_current)
        self.assertEqual(PackListDocument.objects.filter(is_current=True).count(), 1)
        self.assertEqual(PackListDocument.objects.filter(status=PackListDocument.ARCHIVED).count(), 1)

    def test_qc_recheck_does_not_create_duplicate_scan(self):
        rows = self.rows()
        rows[0]['serial_number'] = 'SN-702-003'
        document, _, _ = _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            rows,
            content_hash='a' * 64,
            package_qty=2,
        )
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        first_batch = PackListImportBatch.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
            status=PackListImportBatch.PASSED,
            source_type='UPLOAD',
        )
        record, _ = _scan(
            self.openid,
            self.request(),
            self.asn_code,
            '702-S',
            'SN-702-003',
            damaged=True,
            source='inspection',
            import_batch=first_batch,
        )
        self.assertEqual(record.status, AsnSerialRecord.DAMAGED)
        second_batch = PackListImportBatch.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
            status=PackListImportBatch.PASSED,
            source_type='UPLOAD',
        )
        record, _ = _scan(
            self.openid,
            self.request(),
            self.asn_code,
            '702-S',
            'SN-702-003',
            damaged=False,
            source='inspection',
            import_batch=second_batch,
        )
        self.assertEqual(record.status, AsnSerialRecord.ACCEPTED)
        self.assertEqual(record.scan_count, 0)
        self.assertEqual(_summary(self.openid, self.asn_code)['qc_status'], 'PASSED')

    def test_late_pack_list_sn_mismatch_is_an_open_reconciliation_exception(self):
        original_rows = self.rows()
        original_rows[0]['serial_number'] = 'SN-702-ORIGINAL'
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            original_rows,
            content_hash='a' * 64,
            package_qty=2,
        )
        _scan(self.openid, self.request(), self.asn_code, '702-S', 'SN-702-ORIGINAL', source='inspection')
        late_rows = self.rows()
        late_rows[0]['serial_number'] = 'SN-702-LATE'
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            late_rows,
            content_hash='b' * 64,
            package_qty=2,
            replace=True,
            late_reference=True,
        )
        summary = _summary(self.openid, self.asn_code)
        self.assertEqual(summary['pack_list_serial_mismatch_count'], 2)
        self.assertEqual(summary['reconciliation_status'], 'EXCEPTION')
        self.assertFalse(summary['ready_for_putaway'])

    def test_damaged_receiving_scan_is_open_exception_until_resolved(self):
        rows = self.rows()
        rows[0]['serial_number'] = 'SN-702-002'
        _create_pack_list(
            self.openid,
            self.request(),
            self.asn_code,
            rows,
            content_hash='a' * 64,
            package_qty=2,
        )
        record, _ = _scan(
            self.openid,
            self.request(),
            self.asn_code,
            '702-S',
            'SN-702-002',
            damaged=True,
            row={'note': 'Packaging damaged during receiving'},
        )
        self.assertEqual(record.status, AsnSerialRecord.DAMAGED)
        self.assertEqual(record.note, 'Packaging damaged during receiving')
        self.assertFalse(_summary(self.openid, self.asn_code)['ready_for_putaway'])

    def test_qc_evidence_url_preserves_case(self):
        record, _ = _scan(
            self.openid,
            self.request(),
            self.asn_code,
            '702-S',
            'SN-EVIDENCE-001',
            row={'evidence_url': 'https://drive.google.com/drive/u/0/folders/AbC123'},
        )

        self.assertEqual(record.evidence_url, 'https://drive.google.com/drive/u/0/folders/AbC123')

    def test_open_quantity_exception_is_not_ready_for_putaway(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 1
        detail.goods_shortage_qty = 1
        detail.save(update_fields=['goods_actual_qty', 'goods_shortage_qty'])
        summary = _summary(self.openid, self.asn_code)
        self.assertEqual(summary['total_quantity_exceptions'], 1)
        self.assertFalse(summary['ready_for_putaway'])

    def test_resolved_quantity_exception_can_be_ready_for_putaway(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 1
        detail.goods_shortage_qty = 1
        detail.exception_resolved = True
        detail.save(update_fields=['goods_actual_qty', 'goods_shortage_qty', 'exception_resolved'])
        summary = _summary(self.openid, self.asn_code)
        self.assertEqual(summary['total_quantity_exceptions'], 0)
        self.assertTrue(summary['ready_for_putaway'])

    def test_held_serial_is_not_putaway_eligible_but_qc_can_complete(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-OK-001',
            status=AsnSerialRecord.ACCEPTED,
            is_expected=True,
            is_received=True,
        )
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-HOLD-001',
            status=AsnSerialRecord.DAMAGED,
            is_expected=True,
            is_received=True,
            damaged=True,
            exception_resolved=True,
            exception_resolution_action=HOLD_QUARANTINE,
            exception_resolution_note='Move damaged unit to quarantine.',
            exception_resolution_location='QC-HOLD-01',
        )

        summary = _summary(self.openid, self.asn_code)
        line = summary['lines'][0]

        self.assertTrue(summary['qc_complete'])
        self.assertEqual(summary['total_eligible_for_putaway'], 1)
        self.assertEqual(summary['total_held_serials'], 1)
        self.assertEqual(line['eligible_for_putaway'], 1)
        self.assertTrue(summary['ready_for_putaway'])

    def test_rejected_serials_are_not_putaway_eligible(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 1
        detail.save(update_fields=['goods_actual_qty'])
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-REJECT-001',
            status=AsnSerialRecord.REJECTED,
            is_expected=True,
            is_received=True,
            exception_resolved=True,
            exception_resolution_action=REJECT_RETURN,
            exception_resolution_note='Return damaged unit.',
            exception_resolution_location='RETURN-01',
        )

        summary = _summary(self.openid, self.asn_code)

        self.assertTrue(summary['qc_complete'])
        self.assertEqual(summary['total_rejected_serials'], 1)
        self.assertEqual(summary['total_eligible_for_putaway'], 0)
        self.assertFalse(summary['ready_for_putaway'])

    def test_repair_serial_keeps_partial_putaway_available(self):
        detail = AsnDetailModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        detail.goods_actual_qty = 2
        detail.save(update_fields=['goods_actual_qty'])
        asn = AsnListModel.objects.get(asn_code=self.asn_code, openid=self.openid)
        asn.actual_arrival_at = timezone.now()
        asn.asn_status = 4
        asn.save(update_fields=['actual_arrival_at', 'asn_status'])
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-OK-REPAIR-001',
            status=AsnSerialRecord.ACCEPTED,
            is_expected=True,
            is_received=True,
        )
        AsnSerialRecord.objects.create(
            openid=self.openid,
            asn_code=self.asn_code,
            goods_code='702-S',
            serial_number='SN-REPAIR-001',
            status=AsnSerialRecord.DAMAGED,
            is_expected=True,
            is_received=True,
            damaged=True,
            exception_resolved=True,
            exception_resolution_action=REPAIR_REWORK,
            exception_resolution_note='Needs repair and reinspection.',
            exception_resolution_location='REPAIR-01',
        )

        summary = _summary(self.openid, self.asn_code)
        self.assertTrue(summary['qc_complete'])
        self.assertEqual(summary['total_eligible_for_putaway'], 1)
        self.assertEqual(summary['total_repair_serials'], 1)
        self.assertTrue(summary['ready_for_putaway'])

        data = ASNListGetSerializer(asn, context={}).data
        self.assertEqual(data['serial_acceptance']['repair'], 1)
        self.assertEqual(data['operational_status'], 'READY_FOR_PUTAWAY_PARTIAL')
        self.assertEqual(data['next_action_code'], 'ASSIGN_DRIVER_PUTAWAY')
