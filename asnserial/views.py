from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO

from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import APIException, ValidationError

from asn.models import AsnDetailModel, AsnListModel
from asn.services import asn_detail_reference_errors
from dn.models import DnDetailModel, DnListModel
from binset.models import ListModel as Bin
from stock.models import StockBinModel, StockListModel
from utils.md5 import Md5
from staff.models import ListModel as Staff
from supplier.shortname import generated_supplier_short_name
from receiving.services import assert_legacy_asn_putaway_allowed

from .models import (
    ACCEPT_FOR_PUTAWAY,
    HOLD_QUARANTINE,
    LEGACY_ACCEPT_EXCEPTION,
    NON_PUTAWAY_RESOLUTIONS,
    REPAIR_REWORK,
    REJECT_RETURN,
    AsnSerialRecord,
    ExceptionQuantityMovement,
    PackListDocument,
    PackListImportBatch,
    PackListLine,
    AgentCommandPreview,
    EntityProvenance,
    OperationAudit,
    SourceEvidence,
    SourceIntakeRecord,
    MailboxSyncRun,
    MailboxSyncState,
    PUTAWAY_APPROVED_RESOLUTIONS,
    resolution_allows_putaway,
)
from .agent import (
    AI_AGENT_CLIENT,
    SUPPORTED_OPERATIONS,
    complete_preview,
    consume_preview,
    create_preview,
    is_agent_request,
    is_web_request,
    request_payload,
    require_agent_role,
    agent_roles_for_operation,
    approve_ai_preview,
    approve_web_preview,
    create_source_capture,
    create_web_preview,
    consume_web_preview,
)
from .intake import (
    INTAKE_ROLES,
    _forwarded_value,
    _original_email,
    _original_value,
    ensure_source_intake_record,
    safe_source_metadata,
    update_source_intake,
)
from .permissions import AgentPreviewPermission, SourceIntakePermission


EXCEPTION_STATUSES = {
    AsnSerialRecord.UNEXPECTED,
    AsnSerialRecord.DUPLICATE,
    AsnSerialRecord.WRONG_SKU,
    AsnSerialRecord.DAMAGED,
    AsnSerialRecord.REJECTED,
}

SERIAL_EXCEPTION_ACTIONS = {
    ACCEPT_FOR_PUTAWAY,
    LEGACY_ACCEPT_EXCEPTION,
    HOLD_QUARANTINE,
    REPAIR_REWORK,
    REJECT_RETURN,
    'WAIVE_MISSING',
    'REOPEN',
}


def _resolved_putaway_count(records):
    return records.filter(
        exception_resolved=True,
        exception_resolution_action__in=PUTAWAY_APPROVED_RESOLUTIONS,
    ).count()


def _resolved_hold_count(records):
    return records.filter(
        exception_resolved=True,
        exception_resolution_action=HOLD_QUARANTINE,
    ).count()


def _resolved_reject_count(records):
    return records.filter(
        exception_resolved=True,
        exception_resolution_action=REJECT_RETURN,
    ).count()


def _resolved_repair_count(records):
    return records.filter(
        exception_resolved=True,
        exception_resolution_action=REPAIR_REWORK,
    ).count()


def _openid(request):
    auth = getattr(request, 'auth', None)
    value = getattr(auth, 'openid', None)
    if not value:
        raise APIException({'detail': 'Authentication is required'})
    return str(value)


class AgentCommandPreviewView(APIView):
    """Create a short-lived, tenant-scoped preview token for CLI mutations."""

    permission_classes = [AgentPreviewPermission]

    def post(self, request):
        def reject(detail, **extra):
            payload = {'detail': detail}
            payload.update(extra)
            raise ValidationError(payload)

        operation = _clean(request.data.get('operation')).lower()
        if operation not in SUPPORTED_OPERATIONS:
            reject('Unsupported agent operation', operation=operation)
        require_agent_role(request, agent_roles_for_operation(operation))
        payload = request.data.get('payload') or {}
        if isinstance(payload, str):
            try:
                import json
                payload = json.loads(payload)
            except Exception:
                reject('payload must be a JSON object')
        if not isinstance(payload, dict):
            reject('payload must be a JSON object')
        resource_id = _text(request.data.get('resource_id'))
        asn_code = _clean(request.data.get('asn_code'))
        source_evidence_id = _text(request.data.get('source_evidence_id'))
        if operation in {'asn.eta', 'asn.arrival', 'asn.reserve_staging', 'asn.unload_start', 'asn.unload_finish', 'asn.receive'} and not resource_id:
            reject('resource_id is required for %s' % operation)
        if operation in {'asn.putaway', 'packlist.confirm', 'serial.resolve', 'serial.exception_move'} and not resource_id:
            reject('resource_id is required for %s' % operation)
        if operation in {'asn.receive', 'asn.reserve_staging', 'asn.unload_start', 'asn.unload_finish', 'asn.putaway', 'asn.putaway_bulk', 'serial.resolve_quantity', 'serial.exception_move', 'serial.exception_move_quantity', 'packlist.import', 'serial.import', 'inspection.import'} and not asn_code:
            reject('asn_code is required for %s' % operation)
        if operation in {'asn.eta', 'asn.arrival', 'asn.reserve_staging', 'asn.unload_start', 'asn.unload_finish', 'asn.receive'}:
            asn = AsnListModel.objects.filter(
                openid=request.auth.openid,
                id=resource_id,
                is_delete=False,
            ).first()
            if asn is None:
                reject('ASN does not exist')
            if asn_code and asn.asn_code != asn_code:
                reject('ASN code does not match the selected ASN')
            expected_status = {
                'asn.arrival': 1,
                'asn.reserve_staging': 1,
                'asn.unload_start': 1,
                'asn.unload_finish': 2,
                'asn.receive': 3,
            }.get(operation)
            if expected_status is not None and int(asn.asn_status or 0) != expected_status:
                reject('%s requires ASN status %s' % (operation, expected_status))
            if operation == 'asn.unload_start' and not asn.actual_arrival_at:
                reject('Mark the ASN as arrived before starting unloading')
        if operation == 'asn.eta' and not AsnListModel.objects.filter(
            openid=request.auth.openid, id=resource_id, is_delete=False,
        ).exists():
            reject('ASN does not exist')
        if operation == 'asn.putaway':
            detail = AsnDetailModel.objects.filter(
                openid=request.auth.openid, id=resource_id, is_delete=False,
            ).first()
            if detail is None or detail.asn_status != 4:
                reject('ASN detail is not ready for putaway')
            if asn_code and detail.asn_code != asn_code:
                reject('ASN code does not match the selected ASN detail')
            from asn.views import MoveToBinViewSet
            try:
                with transaction.atomic():
                    validator = MoveToBinViewSet()
                    validator.request = request
                    validator._validate_putaway_request(
                        detail.asn_code,
                        detail,
                        payload.get('qty'),
                        payload.get('bin_name'),
                        payload.get('putaway_driver') or payload.get('driver'),
                    )
            except APIException as exc:
                reject(exc.detail)
        if operation == 'asn.putaway_bulk':
            asn = AsnListModel.objects.filter(
                openid=request.auth.openid,
                asn_code=asn_code,
                asn_status=4,
                is_delete=False,
            ).first()
            if asn is None:
                reject('ASN is not ready for putaway')
        if operation.startswith('outbound.'):
            if operation == 'outbound.create':
                for field in ('customer', 'creater'):
                    if not _clean(payload.get(field)):
                        reject('%s is required' % field)
            elif operation == 'outbound.detail.create':
                dn_code = _clean(payload.get('dn_code'))
                if not dn_code:
                    reject('dn_code is required for outbound.detail.create')
                dn = DnListModel.objects.filter(
                    openid=request.auth.openid,
                    dn_code=dn_code,
                    is_delete=False,
                ).first()
                if dn is None:
                    reject('Delivery note does not exist')
                if int(dn.dn_status or 0) != 1:
                    reject('Outbound detail can only be added to a pre-order delivery note')
                goods_codes = payload.get('goods_code')
                goods_qty = payload.get('goods_qty')
                if not isinstance(goods_codes, list) or not goods_codes:
                    reject('goods_code must be a non-empty list')
                if not isinstance(goods_qty, list) or not goods_qty:
                    reject('goods_qty must be a non-empty list')
                if len(goods_codes) != len(goods_qty):
                    reject('goods_code and goods_qty must have the same length')
            else:
                try:
                    dn_id = int(resource_id)
                except (TypeError, ValueError):
                    reject('resource_id must be a delivery note id')
                dn = DnListModel.objects.filter(
                    openid=request.auth.openid,
                    id=dn_id,
                    is_delete=False,
                ).first()
                if dn is None:
                    reject('Delivery note does not exist')
                expected_status = {
                    'outbound.release': 1,
                    'outbound.order_release': 2,
                    'outbound.pick': 3,
                    'outbound.dispatch': 4,
                    'outbound.pod': 5,
                    'outbound.cancel_intransit': 5,
                }.get(operation)
                if expected_status is not None and int(dn.dn_status or 0) != expected_status:
                    reject('%s requires delivery note status %s' % (operation, expected_status))
        if operation == 'asn.detail.create':
            goods_codes = payload.get('goods_code')
            goods_qty = payload.get('goods_qty')
            if not isinstance(goods_codes, list) or not goods_codes:
                reject('goods_code must be a non-empty list')
            if not isinstance(goods_qty, list) or not goods_qty:
                reject('goods_qty must be a non-empty list')
            if len(goods_codes) != len(goods_qty):
                reject('goods_code and goods_qty must have the same length')
            reference_errors = asn_detail_reference_errors(
                request.auth.openid,
                payload.get('asn_code'),
                payload.get('supplier'),
                goods_codes,
            )
            if reference_errors:
                reject(reference_errors)
        if operation == 'packlist.confirm':
            if not PackListDocument.objects.filter(
                openid=request.auth.openid, id=resource_id, is_current=True,
            ).exists():
                reject('Pack List does not exist')
        if operation == 'serial.resolve':
            if not AsnSerialRecord.objects.filter(openid=request.auth.openid, id=resource_id).exists():
                reject('Serial record does not exist')
        if operation == 'serial.exception_move':
            if not AsnSerialRecord.objects.filter(
                openid=request.auth.openid,
                id=resource_id,
                asn_code=asn_code,
                exception_resolved=True,
                exception_moved=False,
            ).exists():
                reject('Serial exception is not ready for physical movement')
        if operation == 'serial.resolve_quantity':
            if not AsnDetailModel.objects.filter(
                openid=request.auth.openid,
                asn_code=asn_code,
                goods_code=_clean(payload.get('goods_code')),
                is_delete=False,
            ).exists():
                reject('ASN detail does not exist')
        if operation == 'serial.exception_move_quantity':
            if not AsnDetailModel.objects.filter(
                openid=request.auth.openid,
                asn_code=asn_code,
                goods_code=_clean(payload.get('goods_code')),
                exception_resolved=True,
                is_delete=False,
            ).exists():
                reject('Quantity exception is not ready for physical movement')
        return Response(create_preview(
            request,
            operation,
            payload,
            resource_id=resource_id,
            asn_code=asn_code,
            source_evidence_id=source_evidence_id,
        ))


class SourceCaptureView(APIView):
    """Capture source metadata before an AI agent creates an external instruction."""

    def post(self, request):
        sync_run = None
        sync_run_id = request.data.get('sync_run_id')
        if sync_run_id:
            sync_run = MailboxSyncRun.objects.filter(
                id=sync_run_id,
                openid=request.auth.openid,
            ).first()
            if sync_run is None:
                raise ValidationError({'sync_run_id': 'Mailbox sync run does not exist for this tenant'})
        source = create_source_capture(request, request.data)
        intake, created = ensure_source_intake_record(
            source,
            sync_run=sync_run,
            duplicate=bool(getattr(source, '_capture_reused', False)),
        )
        return Response({
            'detail': 'Source evidence captured',
            'duplicate': bool(getattr(source, '_capture_reused', False)),
            'source_evidence': {
                'id': source.id,
                'source_type': source.source_type,
                'operation': source.operation,
                'mailbox_account': source.mailbox_account,
                'message_id': source.message_id,
                'thread_id': source.thread_id,
                'sent_at': source.sent_at.isoformat() if source.sent_at else None,
                'captured_at': source.captured_at.isoformat(),
                'content_hash': source.content_hash,
            },
            'intake_record': {
                'id': intake.id,
                'status': intake.status,
                'operation': intake.operation,
                'document_type': intake.document_type,
                'created': created,
                'next_action': intake.next_action,
            },
        }, status=201)


def _safe_source_metadata(metadata):
    return safe_source_metadata(metadata)


def _safe_source_body(source):
    """Return only an explicitly captured email body, bounded for the UI."""
    metadata = source.metadata if isinstance(source.metadata, dict) else {}
    original = _original_email(metadata)
    for package in (original, metadata):
        for key in ('body', 'text_body', 'email_body'):
            value = package.get(key) if isinstance(package, dict) else None
            if value not in (None, ''):
                return str(value)[:20000]
    return ''


def _source_body_preview(source, limit=240):
    """Return a compact, safe email-body snippet for list-row tooltips."""
    body = ' '.join(_safe_source_body(source).split())
    if len(body) <= limit:
        return body
    return '%s…' % body[:limit - 1].rstrip()


def _email_provenance_payload(source, record):
    metadata = source.metadata if isinstance(source.metadata, dict) else {}
    original = _original_email(metadata)
    original_sent_at = record.sent_at or source.sent_at
    original_payload = {
        'sender_name': _original_value(metadata, 'sender_name', 'from_name') or record.sender_name,
        'sender_email': _original_value(metadata, 'sender_email', 'from_email', 'sender') or record.sender_email,
        'from_raw': str(original.get('from_raw') or '')[:1000],
        'sent_at': original_sent_at,
        'to': original.get('to') or original.get('recipients') or [],
        'cc': original.get('cc') or [],
        'subject': _original_value(metadata, 'subject') or record.subject,
        'message_id': str(original.get('message_id') or '')[:512],
        'thread_id': str(original.get('thread_id') or '')[:512],
        'sent_at_raw': str(original.get('sent_at_raw') or '')[:255],
        'source_location': str(original.get('source_location') or '')[:255],
        'reference': str(original.get('reference') or '')[:255],
    }
    forwarded_payload = {
        'sender_name': str(_forwarded_value(metadata, 'sender_name', 'from_name') or '')[:255],
        'sender_email': str(_forwarded_value(metadata, 'sender_email', 'from_email', 'sender') or '')[:255],
        'subject': str(_forwarded_value(metadata, 'subject') or '')[:1000],
        'sent_at': _forwarded_value(metadata, 'sent_at', 'email_sent_at'),
        'received_at': _forwarded_value(metadata, 'received_at', 'email_received_at'),
        'message_id': str(_forwarded_value(metadata, 'message_id') or source.message_id)[:512],
        'thread_id': str(_forwarded_value(metadata, 'thread_id') or source.thread_id)[:512],
    }
    return original_payload, forwarded_payload


def _intake_payload(record, detail=False):
    source = record.source
    metadata = source.metadata if isinstance(source.metadata, dict) else {}
    original_email, forwarded_email = _email_provenance_payload(source, record)
    payload = {
        'id': record.id,
        'source_evidence_id': source.id,
        'sync_run_id': record.sync_run_id,
        'status': record.status,
        'operation': record.operation,
        'document_type': record.document_type,
        'mailbox_account': record.mailbox_account,
        'sender_name': record.sender_name,
        'sender_email': record.sender_email,
        'subject': record.subject,
        'external_reference': record.external_reference,
        'matched_entity_type': record.matched_entity_type,
        'matched_entity_ref': record.matched_entity_ref,
        'owner_role': record.owner_role,
        'next_action': record.next_action,
        'exception_summary': record.exception_summary,
        'last_error': record.last_error,
        'classification_confidence': record.classification_confidence,
        'sent_at': record.sent_at or source.sent_at,
        'sent_at_raw': str(original_email.get('sent_at_raw') or '')[:255],
        'received_at': record.received_at,
        'received_at_raw': _forwarded_value(metadata, 'received_at', 'email_received_at'),
        'updated_at': record.updated_at,
        'source_type': source.source_type,
        'source_status': source.status,
        'email_body_preview': _source_body_preview(source),
        'message_id': source.message_id,
        'thread_id': source.thread_id,
        'content_hash': source.content_hash,
        'captured_at': source.captured_at,
    }
    if detail:
        payload.update({
            'metadata': _safe_source_metadata(record.metadata),
            'email_body': _safe_source_body(source),
            'original_email': original_email,
            'forwarded_email': forwarded_email,
            'storage_uri': source.storage_uri,
            'storage_size': source.storage_size,
            'extractions': [
                {
                    'field_name': item.field_name,
                    'raw_value': item.raw_value,
                    'normalized_value': item.normalized_value,
                    'source_location': item.source_location,
                    'confidence': item.confidence,
                    'human_confirmed': item.human_confirmed,
                    'used_for_write': item.used_for_write,
                }
                for item in source.extractions.all()
            ],
            'attachments': [
                {
                    'id': item.id,
                    'attachment_name': item.attachment_name,
                    'content_type': item.content_type,
                    'content_hash': item.content_hash,
                    'storage_uri': item.storage_uri,
                    'storage_size': item.storage_size,
                    'security_status': item.security_status,
                    'source_location': item.source_location,
                }
                for item in source.attachments.all()
            ],
            'events': [
                {
                    'id': item.id,
                    'status': item.status,
                    'event_type': item.event_type,
                    'message': item.message,
                    'actor_type': item.actor_type,
                    'actor_name': item.actor_name,
                    'created_at': item.created_at,
                }
                for item in record.events.all()[:100]
            ],
        })
    return payload


class SourceIntakeListView(APIView):
    """Independent source intake board API with bounded pagination."""

    permission_classes = [SourceIntakePermission]

    def get(self, request):
        queryset = SourceIntakeRecord.objects.filter(
            openid=request.auth.openid,
        ).select_related('source', 'sync_run')
        for field in ('status', 'operation', 'document_type', 'mailbox_account'):
            value = str(request.query_params.get(field) or '').strip()
            if value:
                queryset = queryset.filter(**{field: value.upper() if field != 'mailbox_account' else value})
        search = str(request.query_params.get('q') or '').strip()
        if search:
            queryset = queryset.filter(
                Q(subject__icontains=search)
                | Q(sender_email__icontains=search)
                | Q(external_reference__icontains=search)
                | Q(matched_entity_ref__icontains=search)
            )
        try:
            limit = min(max(int(request.query_params.get('limit', 50)), 1), 200)
            offset = max(int(request.query_params.get('offset', 0)), 0)
        except (TypeError, ValueError):
            return Response({'detail': 'limit and offset must be integers'}, status=400)
        total = queryset.count()
        counts = {
            row['status']: row['count']
            for row in queryset.values('status').annotate(count=Count('id'))
        }
        return Response({
            'items': [_intake_payload(record) for record in queryset[offset:offset + limit]],
            'total': total,
            'offset': offset,
            'limit': limit,
            'has_more': offset + limit < total,
            'counts': counts,
        })


class SourceIntakeDetailView(APIView):
    permission_classes = [SourceIntakePermission]

    def get(self, request, pk):
        record = SourceIntakeRecord.objects.filter(
            id=pk,
            openid=request.auth.openid,
        ).select_related('source', 'sync_run').prefetch_related(
            'source__extractions', 'source__attachments', 'events',
        ).first()
        if record is None:
            return Response({'detail': 'Source intake record not found'}, status=404)
        return Response(_intake_payload(record, detail=True))


class SourceIntakeUpdateView(APIView):
    """Update classification/state from Codex without touching WMS business data."""

    permission_classes = [AgentPreviewPermission]

    def post(self, request, pk):
        from .agent import require_agent_role

        require_agent_role(request, INTAKE_ROLES)
        record = SourceIntakeRecord.objects.filter(
            id=pk,
            openid=request.auth.openid,
        ).select_related('source').first()
        if record is None:
            return Response({'detail': 'Source intake record not found'}, status=404)
        actor_type = 'AI_AGENT' if str(request.META.get('HTTP_X_AGENT_CLIENT') or '').lower() == 'greaterwms-ai' else 'CLI'
        actor_name = str(getattr(request.auth, 'staff_name', '') or '')
        updated = update_source_intake(
            record,
            request.data,
            actor_type=actor_type,
            actor_name=actor_name,
        )
        return Response(_intake_payload(updated))


class MailboxSyncRunCreateView(APIView):
    """Start a Codex Automation mailbox scan."""

    permission_classes = [AgentPreviewPermission]

    def post(self, request):
        from .agent import require_agent_role

        require_agent_role(request, INTAKE_ROLES)
        mailbox_account = str(request.data.get('mailbox_account') or '').strip()
        if not mailbox_account:
            raise ValidationError({'mailbox_account': 'mailbox_account is required'})
        trigger_source = str(
            request.data.get('trigger_source') or MailboxSyncRun.CODEX_AUTOMATION
        ).strip().upper()
        if trigger_source not in {MailboxSyncRun.CODEX_AUTOMATION, MailboxSyncRun.MANUAL}:
            raise ValidationError({'trigger_source': 'Use CODEX_AUTOMATION or MANUAL'})
        automation_run_id = str(request.data.get('automation_run_id') or '')[:255]
        metadata = request.data.get('metadata') if isinstance(request.data.get('metadata'), dict) else {}
        with transaction.atomic():
            state, _ = MailboxSyncState.objects.get_or_create(
                openid=request.auth.openid,
                mailbox_account=mailbox_account[:255],
            )
            state = MailboxSyncState.objects.select_for_update().get(id=state.id)
            now = timezone.now()
            if state.active_run_id and state.lease_expires_at and state.lease_expires_at > now:
                active_run = MailboxSyncRun.objects.filter(
                    id=state.active_run_id,
                    status=MailboxSyncRun.RUNNING,
                ).first()
                if active_run is not None:
                    return Response({
                        'detail': 'A mailbox sync is already running for this account',
                        'code': 'MAILBOX_SYNC_IN_PROGRESS',
                        'run_id': active_run.id,
                        'cursor': state.cursor,
                        'lease_expires_at': state.lease_expires_at,
                    }, status=409)
            run = MailboxSyncRun.objects.create(
                openid=request.auth.openid,
                mailbox_account=mailbox_account[:255],
                trigger_source=trigger_source,
                automation_run_id=automation_run_id,
                cursor_before=state.cursor,
                metadata=metadata,
            )
            state.active_run = run
            state.lease_expires_at = now + timedelta(minutes=30)
            state.last_error = ''
            state.save(update_fields=['active_run', 'lease_expires_at', 'last_error', 'updated_at'])
        return Response({
            'id': run.id,
            'status': run.status,
            'mailbox_account': run.mailbox_account,
            'cursor_before': run.cursor_before,
            'lease_expires_at': state.lease_expires_at,
            'started_at': run.started_at,
        }, status=201)


class MailboxSyncStateView(APIView):
    """Return the durable cursor for one tenant mailbox."""

    permission_classes = [AgentPreviewPermission]

    def get(self, request):
        from .agent import require_agent_role

        require_agent_role(request, INTAKE_ROLES)
        mailbox_account = str(request.query_params.get('mailbox_account') or '').strip()
        if not mailbox_account:
            raise ValidationError({'mailbox_account': 'mailbox_account is required'})
        state = MailboxSyncState.objects.filter(
            openid=request.auth.openid,
            mailbox_account=mailbox_account,
        ).select_related('active_run', 'last_successful_run').first()
        if state is None:
            return Response({
                'mailbox_account': mailbox_account,
                'cursor': '',
                'active': False,
            })
        active = bool(
            state.active_run_id
            and state.active_run
            and state.active_run.status == MailboxSyncRun.RUNNING
            and state.lease_expires_at
            and state.lease_expires_at > timezone.now()
        )
        return Response({
            'mailbox_account': state.mailbox_account,
            'cursor': state.cursor,
            'active': active,
            'active_run_id': state.active_run_id if active else None,
            'lease_expires_at': state.lease_expires_at if active else None,
            'last_successful_run_id': state.last_successful_run_id,
            'last_error': state.last_error,
        })


class MailboxSyncRunCompleteView(APIView):
    """Close a Codex Automation mailbox scan with counters and cursor."""

    permission_classes = [AgentPreviewPermission]

    def post(self, request, pk):
        from .agent import require_agent_role

        require_agent_role(request, INTAKE_ROLES)
        with transaction.atomic():
            run = MailboxSyncRun.objects.select_for_update().filter(
                id=pk,
                openid=request.auth.openid,
            ).first()
            if run is None:
                return Response({'detail': 'Mailbox sync run not found'}, status=404)
            if run.status != MailboxSyncRun.RUNNING:
                return Response({
                    'id': run.id,
                    'status': run.status,
                    'cursor_after': run.cursor_after,
                    'detail': 'Mailbox sync run was already completed',
                })
            status = str(request.data.get('status') or MailboxSyncRun.SUCCEEDED).strip().upper()
            if status not in {MailboxSyncRun.SUCCEEDED, MailboxSyncRun.PARTIAL, MailboxSyncRun.FAILED}:
                raise ValidationError({'status': 'Sync run must finish as SUCCEEDED, PARTIAL or FAILED'})
            state = MailboxSyncState.objects.select_for_update().filter(
                openid=run.openid,
                mailbox_account=run.mailbox_account,
            ).first()
            if state is not None and state.active_run_id not in (None, run.id):
                return Response({
                    'detail': 'This sync run no longer owns the mailbox lease',
                    'code': 'MAILBOX_SYNC_LEASE_LOST',
                }, status=409)
            for field in ('fetched_count', 'captured_count', 'duplicate_count', 'review_count', 'failed_count'):
                if field in request.data:
                    try:
                        setattr(run, field, max(int(request.data.get(field)), 0))
                    except (TypeError, ValueError):
                        raise ValidationError({field: 'must be a non-negative integer'})
            run.status = status
            run.cursor_after = str(request.data.get('cursor_after') or '')[:1000]
            run.error_summary = str(request.data.get('error_summary') or '')[:4000]
            run.completed_at = timezone.now()
            run.save(update_fields=[
                'status', 'cursor_after', 'error_summary', 'completed_at',
                'fetched_count', 'captured_count', 'duplicate_count', 'review_count', 'failed_count',
            ])
            if state is not None:
                if status == MailboxSyncRun.SUCCEEDED and run.cursor_after:
                    state.cursor = run.cursor_after
                    state.last_successful_run = run
                state.active_run = None
                state.lease_expires_at = None
                state.last_error = run.error_summary if status != MailboxSyncRun.SUCCEEDED else ''
                state.save(update_fields=[
                    'cursor', 'last_successful_run', 'active_run', 'lease_expires_at',
                    'last_error', 'updated_at',
                ])
        return Response({
            'id': run.id,
            'status': run.status,
            'fetched_count': run.fetched_count,
            'captured_count': run.captured_count,
            'duplicate_count': run.duplicate_count,
            'review_count': run.review_count,
            'failed_count': run.failed_count,
            'state_cursor': state.cursor if state is not None else run.cursor_after,
            'completed_at': run.completed_at,
        })


class SourceEvidenceListView(APIView):
    """Read source evidence and extraction summaries within the current tenant."""

    def get(self, request):
        queryset = SourceEvidence.objects.filter(openid=request.auth.openid)
        operation = _clean(request.query_params.get('operation')).lower()
        # Preserve case-sensitive evidence identifiers. `_clean()` uppercases
        # values for business codes, which would break mailbox/message/hash
        # lookups and make duplicate preflight unreliable.
        mailbox_account = _text(request.query_params.get('mailbox_account'))[:255]
        message_id = _text(request.query_params.get('message_id'))[:512]
        content_hash = _text(request.query_params.get('content_hash'))[:64]
        if operation:
            queryset = queryset.filter(operation=operation)
        if mailbox_account:
            queryset = queryset.filter(mailbox_account=mailbox_account)
        if message_id:
            queryset = queryset.filter(message_id=message_id)
        if content_hash:
            queryset = queryset.filter(content_hash=content_hash)
        results = []
        for source in queryset.prefetch_related('extractions', 'provenance')[:200]:
            results.append({
                'id': source.id,
                'source_type': source.source_type,
                'operation': source.operation,
                'status': source.status,
                'captured_by': source.captured_by_name or source.captured_by,
                'mailbox_account': source.mailbox_account,
                'message_id': source.message_id,
                'thread_id': source.thread_id,
                'sent_at': source.sent_at,
                'captured_at': source.captured_at,
                'content_hash': source.content_hash,
                'metadata': source.metadata,
                'extractions': [
                    {
                        'field_name': item.field_name,
                        'normalized_value': item.normalized_value,
                        'source_location': item.source_location,
                        'confidence': item.confidence,
                        'human_confirmed': item.human_confirmed,
                        'used_for_write': item.used_for_write,
                    }
                    for item in source.extractions.all()
                ],
                'provenance': [
                    {
                        'entity_type': item.entity_type,
                        'entity_ref': item.entity_ref,
                        'field_name': item.field_name,
                        'normalized_value': item.normalized_value,
                        'used_for_write': item.used_for_write,
                    }
                    for item in source.provenance.all()
                ],
            })
        return Response({'count': queryset.count(), 'results': results})


class OperationAuditListView(APIView):
    """Expose safe audit summaries without tokens, passwords or raw payloads."""

    def get(self, request):
        queryset = OperationAudit.objects.filter(openid=request.auth.openid).select_related('source_evidence')
        operation = _clean(request.query_params.get('operation')).lower()
        status = _clean(request.query_params.get('status'))
        if operation:
            queryset = queryset.filter(operation=operation)
        if status:
            queryset = queryset.filter(status=status)
        return Response({
            'count': queryset.count(),
            'results': [
                {
                    'id': audit.id,
                    'operation': audit.operation,
                    'execution_surface': audit.execution_surface,
                    'status': audit.status,
                    'operator_name': audit.operator_name,
                    'operator_role': audit.operator_role,
                    'payload_hash': audit.payload_hash,
                    'source_evidence_id': audit.source_evidence_id,
                    'created_at': audit.created_at,
                    'approved_at': audit.approved_at,
                    'failure_reason': audit.failure_reason,
                    'result': audit.result,
                }
                for audit in queryset[:200]
            ],
        })


class WebWorkflowPreviewView(APIView):
    """Create a no-token Web preview and its automatic WEB_FORM source record."""

    def post(self, request):
        operation = _clean(request.data.get('operation')).lower()
        payload = request.data.get('payload')
        if not isinstance(payload, dict):
            raise ValidationError({'payload': 'A JSON object is required'})
        if not isinstance(payload.get('header'), dict) or not isinstance(payload.get('detail'), dict):
            raise ValidationError({'payload': 'Web preview payload must contain header and detail objects'})
        return Response(create_web_preview(
            request,
            operation,
            payload,
            page=request.data.get('page') or operation,
        ))


class _WorkflowRequest:
    """Small request adapter used to keep Web approval atomic across two legacy writes."""

    def __init__(self, parent, data, stage, surface='WEB'):
        self.data = data
        self.auth = parent.auth
        self.user = parent.user
        self.GET = getattr(parent, 'GET', {})
        self.query_params = getattr(parent, 'query_params', self.GET)
        self.META = dict(parent.META)
        self.META['HTTP_%s_WORKFLOW_STAGE' % surface] = stage
        self.method = 'POST'


def _call_workflow_create(view_class, parent_request, data, stage, surface='WEB'):
    view = view_class()
    view.action = 'create'
    view.format_kwarg = None
    child_request = _WorkflowRequest(parent_request, data, stage, surface=surface)
    view.request = child_request
    return view.create(child_request)


@transaction.atomic
def _execute_external_workflow(request, pk, surface):
    """Approve and execute an ASN/Outbound preview for one locked entry surface."""
    command = approve_ai_preview(request, pk) if surface == 'AI' else approve_web_preview(request, pk)
    if command.status == AgentCommandPreview.EXECUTED:
        return command.result, 200

    payload = command.preview_payload
    header = dict(payload.get('header') or {})
    detail = dict(payload.get('detail') or {})
    control_field = 'agent_preview_id' if surface == 'AI' else 'web_preview_id'
    header[control_field] = command.id
    detail[control_field] = command.id
    header['web_preview_stage'] = 'header'
    detail['web_preview_stage'] = 'detail'

    if command.operation == 'asn.create':
        from asn.views import AsnDetailViewSet, AsnListViewSet
        parent_response = _call_workflow_create(AsnListViewSet, request, header, 'header', surface=surface)
        asn_code = parent_response.data.get('asn_code')
        detail['asn_code'] = asn_code
        detail['supplier'] = detail.get('supplier') or header.get('supplier')
        detail_response = _call_workflow_create(AsnDetailViewSet, request, detail, 'detail', surface=surface)
        result = {
            'detail': 'ASN created',
            'execution_surface': surface.lower(),
            'preview_id': command.id,
            'source_evidence_id': command.source_evidence_id,
            'asn': parent_response.data,
            'asn_detail': detail_response.data,
        }
    elif command.operation == 'outbound.create':
        from dn.views import DnDetailViewSet, DnListViewSet
        parent_response = _call_workflow_create(DnListViewSet, request, header, 'header', surface=surface)
        dn_code = parent_response.data.get('dn_code')
        detail['dn_code'] = dn_code
        detail['customer'] = detail.get('customer') or header.get('customer')
        detail_response = _call_workflow_create(DnDetailViewSet, request, detail, 'detail', surface=surface)
        result = {
            'detail': 'Outbound order created',
            'execution_surface': surface.lower(),
            'preview_id': command.id,
            'source_evidence_id': command.source_evidence_id,
            'outbound': parent_response.data,
            'outbound_detail': detail_response.data,
        }
    else:
        raise ValidationError({'detail': 'Unsupported %s workflow operation' % surface.lower()})

    command.refresh_from_db()
    if command.status != AgentCommandPreview.EXECUTED:
        complete_preview(command, result)
    return result, 201


class AgentCommandApproveView(APIView):
    """Structured AI approval: no CLI token is exposed or accepted."""

    permission_classes = [AgentPreviewPermission]

    def post(self, request, pk):
        if request.META.get('HTTP_X_AGENT_CLIENT', '').strip().lower() != AI_AGENT_CLIENT:
            raise ValidationError({'detail': 'AI approval requires the AI execution surface', 'code': 'AI_SURFACE_REQUIRED'})
        result, status = _execute_external_workflow(request, pk, 'AI')
        return Response(result, status=status)


class WebWorkflowApproveView(APIView):
    """Web button approval: execute the same workflow without a token."""

    @transaction.atomic
    def post(self, request, pk):
        if not is_web_request(request):
            raise ValidationError({'detail': 'Web approval cannot be called from the AI or CLI surface', 'code': 'WEB_SURFACE_REQUIRED'})
        result, status = _execute_external_workflow(request, pk, 'WEB')
        return Response(result, status=status)


def _operator_name(request, openid):
    operator_id = request.META.get('HTTP_OPERATOR')
    staff = None
    if operator_id:
        staff = Staff.objects.filter(openid=openid, id=operator_id, is_delete=False).first()
    return staff.staff_name if staff else str(operator_id or '')


def _clean(value):
    if value is None:
        return ''
    return str(value).strip().upper()


def _text(value):
    """Preserve free text such as evidence URLs without SKU normalization."""
    return str(value or '').strip()


def _is_damage_flag(value):
    normalized = str(value or '').strip().lower()
    return normalized in {'1', 'true', 'yes', 'y', 'ng', 'nok'} or normalized.startswith((
        'damage', 'defect', 'fail', 'reject',
    ))


def _asn_detail(openid, asn_code, goods_code):
    asn = AsnListModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False).first()
    if not asn:
        raise APIException({'detail': 'ASN Code does not exists'})
    detail = AsnDetailModel.objects.filter(
        openid=openid,
        asn_code=asn_code,
        goods_code=goods_code,
        is_delete=False,
    ).first()
    if not detail:
        raise APIException({'detail': 'Goods Code is not part of this ASN'})
    return asn, detail


def _date_value(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _number(value, default=Decimal('0')):
    try:
        return Decimal(str(value or default).replace(',', '').strip())
    except (InvalidOperation, AttributeError):
        return default


def _current_pack_list(openid, asn_code):
    return PackListDocument.objects.filter(
        openid=openid,
        asn_code=asn_code,
        is_current=True,
        status=PackListDocument.CONFIRMED,
    ).first()


def _pack_list_json(document):
    serials = document.serial_records.all()
    lines = document.lines.filter(is_current=True) if document.is_current else document.lines.all()
    return {
        'id': document.id,
        'asn_code': document.asn_code,
        'version': document.version,
        'source_type': document.source_type,
        'status': document.status,
        'is_current': document.is_current,
        'late_reference': document.late_reference,
        'has_serials': document.has_serials,
        'package_qty': document.package_qty,
        'note': document.note,
        'created_by': document.created_by,
        'confirmed_by': document.confirmed_by,
        'confirmed_at': document.confirmed_at.isoformat() if document.confirmed_at else None,
        'create_time': document.create_time.isoformat() if document.create_time else None,
        'line_count': lines.count(),
        'total_qty': sum(line.goods_qty for line in lines),
        'lines': [
            {
                'goods_code': line.goods_code,
                'customer_goods_code': line.customer_goods_code,
                'customer_ssku': line.customer_ssku,
                'goods_qty': line.goods_qty,
                'total_qty': line.total_qty,
                'package_type': line.package_type,
                'goods_desc': line.goods_desc,
                'source_row': line.source_row,
            }
            for line in lines
        ],
        'expected_serial_count': serials.filter(is_expected=True).count(),
        'received_serial_count': serials.filter(is_received=True).count(),
    }


def _inspection_batch_json(batch):
    return {
        'id': batch.id,
        'asn_code': batch.asn_code,
        'import_type': batch.import_type,
        'status': batch.status,
        'source_type': batch.source_type,
        'row_count': batch.row_count,
        'matched_count': batch.matched_count,
        'accepted_count': batch.accepted_count,
        'exception_count': batch.exception_count,
        'note': batch.note,
        'evidence_url': batch.evidence_url,
        'imported_by': batch.imported_by,
        'created_at': batch.created_at.isoformat() if batch.created_at else None,
    }


def _receiving_started(openid, asn_code):
    if AsnListModel.objects.filter(
        openid=openid,
        asn_code=asn_code,
        is_delete=False,
        asn_status__gte=3,
    ).exists():
        return True
    if AsnSerialRecord.objects.filter(openid=openid, asn_code=asn_code, is_received=True).exists():
        return True
    return AsnDetailModel.objects.filter(
        openid=openid,
        asn_code=asn_code,
        is_delete=False,
        goods_actual_qty__gt=0,
    ).exists()


def _pack_list_serial_mismatch(document, records):
    if not document or not document.has_serials:
        return {'total': 0, 'missing': [], 'unexpected': [], 'wrong_sku': [], 'by_goods': {}}
    expected = {
        record.serial_number: record.goods_code
        for record in document.serial_records.filter(is_expected=True)
    }
    received = {
        record.serial_number: record.goods_code
        for record in records
        if record.is_received
    }
    missing = sorted(set(expected) - set(received))
    unexpected = sorted(set(received) - set(expected))
    wrong_sku = sorted(
        serial_number for serial_number in set(expected).intersection(received)
        if _clean(expected[serial_number]) != _clean(received[serial_number])
    )
    by_goods = {}
    for serial_number in missing:
        goods_code = expected[serial_number]
        by_goods[goods_code] = by_goods.get(goods_code, 0) + 1
    for serial_number in unexpected:
        goods_code = received[serial_number]
        by_goods[goods_code] = by_goods.get(goods_code, 0) + 1
    for serial_number in wrong_sku:
        goods_code = expected[serial_number]
        by_goods[goods_code] = by_goods.get(goods_code, 0) + 1
    return {
        'total': len(missing) + len(unexpected) + len(wrong_sku),
        'missing': missing,
        'unexpected': unexpected,
        'wrong_sku': wrong_sku,
        'by_goods': by_goods,
    }


def _record_json(record):
    return {
        'id': record.id,
        'asn_code': record.asn_code,
        'goods_code': record.goods_code,
        'expected_goods_code': record.expected_goods_code,
        'scanned_goods_code': record.scanned_goods_code,
        'serial_number': record.serial_number,
        'double_scan_sn': record.double_scan_sn,
        'inbound_po': record.inbound_po,
        'inbound_date': record.inbound_date.isoformat() if record.inbound_date else None,
        'source_location': record.source_location,
        'shipout_ref': record.shipout_ref,
        'source_row': record.source_row,
        'status': record.status,
        'is_expected': record.is_expected,
        'is_received': record.is_received,
        'scan_count': record.scan_count,
        'damaged': record.damaged,
        'note': record.note,
        'evidence_url': record.evidence_url,
        'exception_resolved': record.exception_resolved,
        'exception_resolution_action': record.exception_resolution_action,
        'exception_resolution_note': record.exception_resolution_note,
        'resolution_location': record.exception_resolution_location,
        'exception_resolved_by': record.exception_resolved_by,
        'exception_resolved_at': record.exception_resolved_at.isoformat() if record.exception_resolved_at else None,
        'exception_moved': record.exception_moved,
        'exception_move_bin': record.exception_move_bin,
        'exception_moved_at': record.exception_moved_at.isoformat() if record.exception_moved_at else None,
        'expected_by': record.expected_by,
        'received_by': record.received_by,
        'expected_at': record.expected_at.isoformat() if record.expected_at else None,
        'received_at': record.received_at.isoformat() if record.received_at else None,
        'pack_list_id': record.pack_list_id,
    }


def _reconciliation_rows(document, details, records, strict_serial_check, exception_statuses, serial_mismatch=None):
    """Join the customer Pack List, ASN receipt quantities, and QC scan results by SKU."""
    pack_lines = {}
    if document:
        for line in document.lines.filter(is_current=True):
            key = _clean(line.goods_code)
            item = pack_lines.setdefault(key, {
                'customer_goods_codes': set(),
                'customer_sskus': set(),
                'pack_list_qty': 0,
            })
            if line.customer_goods_code:
                item['customer_goods_codes'].add(line.customer_goods_code)
            if line.customer_ssku:
                item['customer_sskus'].add(line.customer_ssku)
            item['pack_list_qty'] += int(line.goods_qty or 0)

    detail_list = list(details)
    detail_map = {_clean(detail.goods_code): detail for detail in detail_list}
    record_groups = {}
    for record in records:
        record_groups.setdefault(_clean(record.goods_code), []).append(record)

    rows = []
    keys = list(detail_map.keys())
    keys.extend(key for key in pack_lines if key not in detail_map)
    for key in keys:
        detail = detail_map.get(key)
        line = pack_lines.get(key, {})
        line_records = record_groups.get(key, [])
        expected_count = sum(1 for record in line_records if record.is_expected)
        accepted_count = sum(1 for record in line_records if record.status == AsnSerialRecord.ACCEPTED)
        resolved_count = sum(1 for record in line_records if record.exception_resolved)
        putaway_eligible_count = sum(
            1 for record in line_records
            if record.status == AsnSerialRecord.ACCEPTED or (
                record.exception_resolved and resolution_allows_putaway(record.exception_resolution_action)
            )
        )
        exception_count = sum(
            1 for record in line_records
            if record.status in exception_statuses and not record.exception_resolved
        )
        serial_mismatch_count = int((serial_mismatch or {}).get('by_goods', {}).get(key, 0))
        received_qty = int(detail.goods_actual_qty or 0) if detail else 0
        planned_qty = int(detail.goods_qty or 0) if detail else 0
        pack_list_qty = int(line.get('pack_list_qty') or 0)
        baseline_qty = pack_list_qty if document else planned_qty
        quantity_exception_qty = 0
        quantity_exception_resolved = False
        if detail:
            quantity_exception_qty = 0 if detail.exception_resolved else (
                int(detail.goods_shortage_qty or 0)
                + int(detail.goods_more_qty or 0)
                + int(detail.goods_damage_qty or 0)
            )
            quantity_exception_resolved = bool(detail.exception_resolved)

        open_exception_count = exception_count + serial_mismatch_count + int(quantity_exception_qty > 0)
        resolved_exception_total = resolved_count + int(quantity_exception_resolved)
        variance = received_qty - baseline_qty
        accepted_qty = putaway_eligible_count if (strict_serial_check or (document and document.has_serials)) else received_qty
        if open_exception_count or variance:
            result = 'EXCEPTION'
        elif not document or document.status == PackListDocument.PENDING:
            result = 'REVIEW'
        elif resolved_exception_total:
            result = 'RESOLVED'
        else:
            result = 'PASSED'

        rows.append({
            'goods_code': detail.goods_code if detail else key,
            'customer_goods_code': ', '.join(sorted(line.get('customer_goods_codes', set()))),
            'customer_ssku': ', '.join(sorted(line.get('customer_sskus', set()))),
            'pack_list_qty': pack_list_qty,
            'asn_qty': planned_qty,
            'received_qty': received_qty,
            'accepted_qty': accepted_qty,
            'putaway_eligible_qty': putaway_eligible_count,
            'variance': variance,
            'baseline': 'PACK_LIST' if document else 'ASN',
            'expected_serial_count': expected_count,
            'customer_sn_status': 'PROVIDED' if document and document.has_serials else 'NOT_PROVIDED',
            'open_exception_count': open_exception_count,
            'resolved_exception_count': resolved_exception_total,
            'quantity_exception_qty': quantity_exception_qty,
            'goods_shortage_qty': int(detail.goods_shortage_qty or 0),
            'goods_more_qty': int(detail.goods_more_qty or 0),
            'goods_damage_qty': int(detail.goods_damage_qty or 0),
            'repair_count': sum(
                1 for record in line_records
                if record.exception_resolved and record.exception_resolution_action == REPAIR_REWORK
            ),
            'serial_mismatch_count': serial_mismatch_count,
            'result': result,
        })
    return rows


def _summary(openid, asn_code):
    details = AsnDetailModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False)
    records = AsnSerialRecord.objects.filter(openid=openid, asn_code=asn_code)
    pack_lists = PackListDocument.objects.filter(openid=openid, asn_code=asn_code, is_current=True)
    current_pack_list = _current_pack_list(openid, asn_code)
    pending_pack_list = pack_lists.filter(status=PackListDocument.PENDING).first()
    active_pack_list = pack_lists.order_by('-version', '-id').first()
    asn = AsnListModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False).first()
    has_expected_serials = records.filter(is_expected=True).exists()
    if current_pack_list and current_pack_list.has_serials:
        verification_mode = 'PACK_LIST'
    elif pending_pack_list:
        verification_mode = 'PACK_LIST_PENDING' if pending_pack_list.has_serials else 'PACK_LIST_QTY'
    elif has_expected_serials:
        verification_mode = 'MANUAL_SN'
    elif current_pack_list:
        verification_mode = 'PACK_LIST_QTY'
    else:
        verification_mode = 'ASN_ONLY'
    strict_serial_check = has_expected_serials
    exception_statuses = EXCEPTION_STATUSES if strict_serial_check else EXCEPTION_STATUSES - {AsnSerialRecord.UNEXPECTED}
    lines = []
    for detail in details:
        line_records = records.filter(goods_code=detail.goods_code)
        expected_count = line_records.filter(is_expected=True).count()
        received_count = line_records.filter(is_received=True).count()
        accepted_count = line_records.filter(status=AsnSerialRecord.ACCEPTED).count()
        resolved_count = line_records.filter(exception_resolved=True).count()
        resolved_putaway_count = _resolved_putaway_count(line_records)
        exception_count = line_records.filter(status__in=exception_statuses, exception_resolved=False).count()
        missing_count = line_records.filter(is_expected=True, is_received=False, exception_resolved=False).count()
        actual_received_qty = int(detail.goods_actual_qty or 0)
        quantity_exception_qty = (
            int(detail.goods_shortage_qty or 0)
            + int(detail.goods_more_qty or 0)
            + int(detail.goods_damage_qty or 0)
        )
        quantity_only_resolved = (
            actual_received_qty > 0
            and not line_records.exists()
            and not strict_serial_check
            and (bool(detail.exception_resolved) or quantity_exception_qty == 0)
        )
        accepted_for_putaway = actual_received_qty if quantity_only_resolved else min(
            accepted_count + resolved_putaway_count,
            actual_received_qty,
        )
        quantity_exception_qty = 0 if detail.exception_resolved else quantity_exception_qty
        lines.append({
            'goods_code': detail.goods_code,
            'planned_qty': detail.goods_qty,
            'received_qty': actual_received_qty,
            'expected_serial_count': expected_count,
            'received_serial_count': received_count,
            'extra_scan_count': max(received_count - actual_received_qty, 0),
            'accepted_serial_count': accepted_count,
            'accepted_for_putaway': accepted_for_putaway,
            'eligible_for_putaway': accepted_for_putaway,
            'resolved_exception_count': resolved_count,
            'held_count': _resolved_hold_count(line_records),
            'repair_count': _resolved_repair_count(line_records),
            'rejected_count': _resolved_reject_count(line_records),
            'missing_serial_count': missing_count,
            'exception_count': exception_count,
            'quantity_exception_qty': quantity_exception_qty,
            'quantity_exception_resolved': bool(detail.exception_resolved),
            'exception_resolved': bool(detail.exception_resolved),
            'exception_resolution_action': detail.exception_resolution_action,
            'exception_resolution_note': detail.exception_resolution_note,
            'resolution_location': detail.exception_resolution_location,
            'ready_for_putaway': (
                quantity_exception_qty == 0 and (
                    not line_records.exists()
                    or (not strict_serial_check and exception_count == 0)
                    or (strict_serial_check and missing_count == 0 and exception_count == 0 and accepted_for_putaway >= detail.goods_actual_qty)
                )
            ),
        })
    exception_total = records.filter(status__in=exception_statuses, exception_resolved=False).count()
    missing_total = records.filter(is_expected=True, is_received=False, exception_resolved=False).count()
    resolved_total = records.filter(exception_resolved=True).count()
    accepted_total = records.filter(status=AsnSerialRecord.ACCEPTED).count()
    actual_received_qty = sum(int(detail.goods_actual_qty or 0) for detail in details)
    physical_putaway_qty = sum(int(detail.sorted_qty or 0) for detail in details)
    scanned_record_count = records.filter(is_received=True).count()
    resolved_putaway_total = _resolved_putaway_count(records)
    quantity_only_resolved = (
        actual_received_qty > 0
        and not records.exists()
        and not has_expected_serials
        and not any(
            not detail.exception_resolved and (
                int(detail.goods_shortage_qty or 0)
                + int(detail.goods_more_qty or 0)
                + int(detail.goods_damage_qty or 0)
            ) > 0
            for detail in details
        )
    )
    accepted_for_putaway_total = actual_received_qty if quantity_only_resolved else min(
        accepted_total + resolved_putaway_total,
        actual_received_qty,
    )
    extra_scan_record_count = max(scanned_record_count - actual_received_qty, 0)
    quantity_exception_total = sum(
        0 if detail.exception_resolved else (
            int(detail.goods_shortage_qty or 0)
            + int(detail.goods_more_qty or 0)
            + int(detail.goods_damage_qty or 0)
        )
        for detail in details
    )
    reconciliation_rows = _reconciliation_rows(
        active_pack_list,
        details,
        records,
        strict_serial_check,
        exception_statuses,
        serial_mismatch=_pack_list_serial_mismatch(active_pack_list, records),
    )
    open_reconciliation_exceptions = sum(row['open_exception_count'] for row in reconciliation_rows)
    resolved_reconciliation_exceptions = sum(row['resolved_exception_count'] for row in reconciliation_rows)
    pack_list_variance = sum(
        abs(int(row['variance'] or 0))
        for row in reconciliation_rows
    ) if active_pack_list else 0
    pack_list_serial_mismatch = _pack_list_serial_mismatch(active_pack_list, records)
    # Quantity-only receiving has no SN rows by design. It is complete here
    # only after all quantity exceptions have been explicitly resolved.
    has_receiving_result = records.exists() or actual_received_qty == 0 or quantity_only_resolved
    qc_complete = bool(
        has_receiving_result
        and not open_reconciliation_exceptions
        and not pack_list_variance
        and missing_total == 0
        and quantity_exception_total == 0
    )
    if open_reconciliation_exceptions or pack_list_variance:
        reconciliation_status = 'EXCEPTION'
    elif not active_pack_list or active_pack_list.status == PackListDocument.PENDING:
        reconciliation_status = 'REVIEW'
    elif resolved_reconciliation_exceptions:
        reconciliation_status = 'RESOLVED'
    else:
        reconciliation_status = 'PASSED'
    receiving_status = 'EXCEPTION' if (open_reconciliation_exceptions or pack_list_variance) else (
        'RESOLVED' if resolved_reconciliation_exceptions else 'PASSED'
    )
    inspection_batches = list(PackListImportBatch.objects.filter(
        openid=openid,
        asn_code=asn_code,
        import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
    )[:10])
    latest_inspection = inspection_batches[0] if inspection_batches else None
    qc_import_incomplete = bool(
        latest_inspection and latest_inspection.status in (
            PackListImportBatch.IMPORTED,
            PackListImportBatch.PARTIAL,
        )
    )
    if qc_import_incomplete:
        qc_complete = False
        if not open_reconciliation_exceptions and not pack_list_variance:
            receiving_status = 'REVIEW'
    if open_reconciliation_exceptions or pack_list_variance or (inspection_batches and missing_total):
        qc_status = 'EXCEPTION'
        reconciliation_status = 'EXCEPTION'
        receiving_status = 'EXCEPTION'
    elif not inspection_batches:
        qc_status = 'NOT_STARTED'
    elif latest_inspection.status in (
        PackListImportBatch.IMPORTED,
        PackListImportBatch.PARTIAL,
    ):
        qc_status = 'PARTIAL'
    else:
        qc_status = 'PASSED'
    all_pack_lists = PackListDocument.objects.filter(
        openid=openid,
        asn_code=asn_code,
    ).order_by('-version', '-id')
    pack_list_status = (
        PackListDocument.CONFIRMED if current_pack_list else
        PackListDocument.PENDING if pending_pack_list else
        'NOT_RECEIVED'
    )
    if active_pack_list and active_pack_list.late_reference:
        pack_list_status = 'LATE' if active_pack_list.status == PackListDocument.CONFIRMED else 'LATE_PENDING'
    return {
        'asn_code': asn_code,
        'customer': asn.supplier if asn else '',
        'customer_short_name': generated_supplier_short_name(asn.supplier if asn else ''),
        'expected_arrival_at': asn.expected_arrival_at.isoformat() if asn and asn.expected_arrival_at else None,
        'actual_arrival_at': asn.actual_arrival_at.isoformat() if asn and asn.actual_arrival_at else None,
        'pack_list_present': pack_lists.exists(),
        'pack_list_status': pack_list_status,
        'pack_list_timing': 'LATE_REFERENCE' if active_pack_list and active_pack_list.late_reference else (
            'BEFORE_RECEIPT' if active_pack_list else 'NOT_RECEIVED'
        ),
        'pack_list_confirmed': bool(current_pack_list),
        'pack_list_has_serials': bool(current_pack_list and current_pack_list.has_serials),
        'active_pack_list': _pack_list_json(active_pack_list) if active_pack_list else None,
        'customer_sn_status': 'PROVIDED' if active_pack_list and active_pack_list.has_serials else 'NOT_PROVIDED',
        'current_pack_list': _pack_list_json(current_pack_list) if current_pack_list else None,
        'pack_list_history': [_pack_list_json(document) for document in all_pack_lists],
        'inspection_batches': [_inspection_batch_json(batch) for batch in inspection_batches],
        'latest_inspection_batch': _inspection_batch_json(latest_inspection) if latest_inspection else None,
        'qc_status': qc_status,
        'qc_complete': qc_complete,
        'qc_import_incomplete': qc_import_incomplete,
        'verification_mode': verification_mode,
        'verification_note': (
            'Receiving scans are not checked against a Pack List yet.'
            if verification_mode == 'ASN_ONLY' else
            'Pack List has quantities only; physical scans are recorded without SN validation.'
            if verification_mode == 'PACK_LIST_QTY' else
            'Pack List with expected SN is pending confirmation.'
            if verification_mode == 'PACK_LIST_PENDING' else
            'Expected SN comes from a Pack List.'
            if verification_mode == 'PACK_LIST' else
            'Expected SN was entered manually.'
        ),
        'lines': lines,
        'reconciliation_status': reconciliation_status,
        'reconciliation_rows': reconciliation_rows,
        'receiving_summary': {
            'expected': records.filter(is_expected=True).count(),
            'received_qty': actual_received_qty,
            'scanned': scanned_record_count,
            'scan_record_count': scanned_record_count,
            'accepted': accepted_total,
            'accepted_for_putaway': accepted_for_putaway_total,
            'eligible_for_putaway': accepted_for_putaway_total,
            'putaway_qty': physical_putaway_qty,
            'held_qty': _resolved_hold_count(records),
            'repair_qty': _resolved_repair_count(records),
            'rejected_qty': _resolved_reject_count(records),
            'extra_scan_records': extra_scan_record_count,
            'open_exceptions': open_reconciliation_exceptions,
            'resolved_exceptions': resolved_reconciliation_exceptions,
            'status': receiving_status,
            'qc_status': qc_status,
            'latest_batch_id': latest_inspection.id if latest_inspection else None,
            'pack_list_variance': pack_list_variance,
        },
        'total_expected_serials': records.filter(is_expected=True).count(),
        'total_received_qty': actual_received_qty,
        'total_received_serials': scanned_record_count,
        'total_scan_records': scanned_record_count,
        'total_accepted_serials': accepted_total,
        'total_resolved_exceptions': resolved_total,
        'total_accepted_for_putaway': accepted_for_putaway_total,
        'total_eligible_for_putaway': accepted_for_putaway_total,
        'total_held_serials': _resolved_hold_count(records),
        'total_repair_serials': _resolved_repair_count(records),
        'total_rejected_serials': _resolved_reject_count(records),
        'total_putaway_qty': physical_putaway_qty,
        'total_extra_scan_records': extra_scan_record_count,
        'total_exception_serials': exception_total,
        'total_missing_serials': missing_total,
        'total_quantity_exceptions': quantity_exception_total,
        'pack_list_variance': pack_list_variance,
        'pack_list_serial_mismatch': pack_list_serial_mismatch,
        'pack_list_serial_mismatch_count': pack_list_serial_mismatch['total'],
        'ready_for_putaway': bool(qc_complete and accepted_for_putaway_total > physical_putaway_qty),
    }


def _save_expected(openid, request, asn_code, goods_code, serial_number, row=None, source='manual', pack_list=None, import_batch=None):
    serial_number = _clean(serial_number)
    if not serial_number:
        raise APIException({'detail': 'Serial Number is required'})
    _, detail = _asn_detail(openid, asn_code, goods_code)
    record = AsnSerialRecord.objects.filter(
        openid=openid,
        asn_code=asn_code,
        serial_number=serial_number,
    ).first()
    now = timezone.now()
    metadata = row or {}
    if record:
        if record.goods_code != goods_code:
            raise APIException({'detail': 'Serial Number already belongs to another SKU in this ASN'})
        record.is_expected = True
        record.expected_goods_code = goods_code
        record.double_scan_sn = _clean(metadata.get('double_scan_sn')) or record.double_scan_sn
        record.inbound_po = _clean(metadata.get('inbound_po')) or record.inbound_po
        record.inbound_date = _date_value(metadata.get('inbound_date')) or record.inbound_date
        record.source_location = _clean(metadata.get('source_location')) or record.source_location
        record.shipout_ref = _clean(metadata.get('shipout_ref')) or record.shipout_ref
        record.source_row = int(metadata.get('source_row') or record.source_row or 0)
        record.note = str(metadata.get('note') or record.note or '').strip()
        record.evidence_url = _text(metadata.get('evidence_url')) or record.evidence_url
        record.pack_list = pack_list or record.pack_list
        record.import_batch = import_batch or record.import_batch
        record.expected_by = _operator_name(request, openid)
        record.expected_at = record.expected_at or now
        if record.is_received and record.scanned_goods_code == goods_code and record.status not in EXCEPTION_STATUSES:
            record.status = AsnSerialRecord.ACCEPTED
            record.exception_resolved = False
            record.exception_resolution_action = ''
            record.exception_resolution_note = ''
            record.exception_resolution_location = ''
            record.exception_resolved_by = ''
            record.exception_resolved_at = None
        record.save()
        return record, False
    expected_count = AsnSerialRecord.objects.filter(
        openid=openid,
        asn_code=asn_code,
        goods_code=goods_code,
        is_expected=True,
    ).count()
    if expected_count >= int(detail.goods_qty):
        raise APIException({'detail': 'Expected SN quantity cannot exceed ASN quantity'})
    record = AsnSerialRecord.objects.create(
        openid=openid,
        asn_code=asn_code,
        goods_code=goods_code,
        expected_goods_code=goods_code,
        serial_number=serial_number,
        double_scan_sn=_clean(metadata.get('double_scan_sn')),
        inbound_po=_clean(metadata.get('inbound_po')),
        inbound_date=_date_value(metadata.get('inbound_date')),
        source_location=_clean(metadata.get('source_location')),
        shipout_ref=_clean(metadata.get('shipout_ref')),
        source_row=int(metadata.get('source_row') or 0),
        note=str(metadata.get('note') or '').strip(),
        evidence_url=_text(metadata.get('evidence_url')),
        import_batch=import_batch,
        status=AsnSerialRecord.EXPECTED,
        is_expected=True,
        expected_by=_operator_name(request, openid),
        expected_at=now,
        pack_list=pack_list,
    )
    return record, True


def _scan_status_without_pack_list(openid, asn_code):
    pack_list = _current_pack_list(openid, asn_code)
    return AsnSerialRecord.UNEXPECTED if pack_list and pack_list.has_serials else AsnSerialRecord.UNVERIFIED


def _scan(openid, request, asn_code, goods_code, serial_number, damaged=False, row=None, source='manual', import_batch=None):
    serial_number = _clean(serial_number)
    goods_code = _clean(goods_code)
    if not serial_number or not goods_code:
        raise APIException({'detail': 'Goods Code and Serial Number are required'})
    _asn_detail(openid, asn_code, goods_code)
    record = AsnSerialRecord.objects.filter(
        openid=openid,
        asn_code=asn_code,
        serial_number=serial_number,
    ).first()
    now = timezone.now()
    metadata = row or {}
    inspection = source in ('inspection', 'qc')
    if record:
        # An inspection workbook is a result snapshot, not another physical scan.
        # Re-importing a later QC round must not turn a valid SN into a duplicate.
        if not inspection:
            record.scan_count += 1
        record.is_received = True
        record.received_at = now
        record.received_by = _operator_name(request, openid)
        record.scanned_goods_code = goods_code
        record.double_scan_sn = _clean(metadata.get('double_scan_sn')) or record.double_scan_sn
        record.inbound_po = _clean(metadata.get('inbound_po')) or record.inbound_po
        record.inbound_date = _date_value(metadata.get('inbound_date')) or record.inbound_date
        record.source_location = _clean(metadata.get('source_location')) or record.source_location
        record.shipout_ref = _clean(metadata.get('shipout_ref')) or record.shipout_ref
        record.source_row = int(metadata.get('source_row') or record.source_row or 0)
        record.import_batch = import_batch or record.import_batch
        record.note = str(metadata.get('note') or record.note or '').strip()
        record.evidence_url = _text(metadata.get('evidence_url')) or record.evidence_url
        record.damaged = bool(damaged) if inspection else record.damaged or bool(damaged)
        if not inspection and record.scan_count > 1:
            record.status = AsnSerialRecord.DUPLICATE
        elif record.goods_code != goods_code:
            record.status = AsnSerialRecord.WRONG_SKU
        elif record.damaged:
            record.status = AsnSerialRecord.DAMAGED
        elif record.is_expected:
            record.status = AsnSerialRecord.ACCEPTED
        else:
            record.status = _scan_status_without_pack_list(openid, asn_code)
        if record.status in EXCEPTION_STATUSES:
            record.exception_resolved = False
            record.exception_resolution_action = ''
            record.exception_resolution_note = ''
            record.exception_resolution_location = ''
            record.exception_resolved_by = ''
            record.exception_resolved_at = None
        elif inspection:
            record.exception_resolved = False
            record.exception_resolution_action = ''
            record.exception_resolution_note = ''
            record.exception_resolution_location = ''
            record.exception_resolved_by = ''
            record.exception_resolved_at = None
        record.save()
        return record, False
    record = AsnSerialRecord.objects.create(
        openid=openid,
        asn_code=asn_code,
        goods_code=goods_code,
        scanned_goods_code=goods_code,
        serial_number=serial_number,
        double_scan_sn=_clean(metadata.get('double_scan_sn')),
        inbound_po=_clean(metadata.get('inbound_po')),
        inbound_date=_date_value(metadata.get('inbound_date')),
        source_location=_clean(metadata.get('source_location')),
        shipout_ref=_clean(metadata.get('shipout_ref')),
        source_row=int(metadata.get('source_row') or 0),
        note=str(metadata.get('note') or '').strip(),
        evidence_url=_text(metadata.get('evidence_url')),
        import_batch=metadata.get('import_batch'),
        status=AsnSerialRecord.DAMAGED if damaged else _scan_status_without_pack_list(openid, asn_code),
        is_expected=False,
        is_received=True,
        scan_count=1,
        damaged=bool(damaged),
        received_by=_operator_name(request, openid),
        received_at=now,
    )
    return record, True


class SerialRecordsView(APIView):
    def get(self, request):
        openid = _openid(request)
        asn_code = _clean(request.query_params.get('asn_code'))
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        records = AsnSerialRecord.objects.filter(openid=openid, asn_code=asn_code)
        goods_code = _clean(request.query_params.get('goods_code'))
        status = _clean(request.query_params.get('status'))
        if goods_code:
            records = records.filter(goods_code=goods_code)
        if status:
            records = records.filter(status=status)
        limit = min(max(int(request.query_params.get('limit', 500)), 1), 5000)
        return Response({'count': records.count(), 'results': [_record_json(r) for r in records[:limit]]})


def _resolution_note(data):
    return str(data.get('note') or data.get('resolution_note') or '').strip()


class SerialExceptionsView(APIView):
    """List current SN and quantity exceptions for QC follow-up."""

    def get(self, request):
        openid = _openid(request)
        asn_code = _clean(request.query_params.get('asn_code'))
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        records = AsnSerialRecord.objects.filter(openid=openid, asn_code=asn_code)
        exception_statuses = EXCEPTION_STATUSES
        results = []
        for record in records.filter(status__in=exception_statuses) | records.filter(is_expected=True, is_received=False):
            if record.exception_resolved:
                continue
            kind = {
                AsnSerialRecord.UNEXPECTED: 'UNEXPECTED_SN',
                AsnSerialRecord.DUPLICATE: 'DUPLICATE_SN',
                AsnSerialRecord.WRONG_SKU: 'WRONG_SKU',
                AsnSerialRecord.DAMAGED: 'DAMAGED_SN',
                AsnSerialRecord.REJECTED: 'REJECTED_SN',
            }.get(record.status, 'MISSING_SN')
            results.append({
                'type': 'SERIAL',
                'id': record.id,
                'asn_code': record.asn_code,
                'goods_code': record.goods_code,
                'serial_number': record.serial_number,
                'kind': kind,
                'status': record.status,
                'quantity': 1,
                'exception_resolved': record.exception_resolved,
                'note': record.note,
                'evidence_url': record.evidence_url,
                'resolution_location': record.exception_resolution_location,
            })

        details = AsnDetailModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False)
        for detail in details:
            quantity = int(detail.goods_shortage_qty or 0) + int(detail.goods_more_qty or 0) + int(detail.goods_damage_qty or 0)
            if quantity <= 0 or detail.exception_resolved:
                continue
            if detail.goods_shortage_qty:
                kind = 'SHORTAGE'
                quantity = int(detail.goods_shortage_qty)
            elif detail.goods_more_qty:
                kind = 'OVERAGE'
                quantity = int(detail.goods_more_qty)
            else:
                kind = 'DAMAGED_QTY'
                quantity = int(detail.goods_damage_qty)
            results.append({
                'type': 'QUANTITY',
                'id': detail.id,
                'asn_code': detail.asn_code,
                'goods_code': detail.goods_code,
                'serial_number': '',
                'kind': kind,
                'status': 'OPEN',
                'quantity': quantity,
                'exception_resolved': detail.exception_resolved,
                'note': detail.exception_resolution_note,
                'resolution_location': detail.exception_resolution_location,
            })
        return Response({'count': len(results), 'results': results})


class SerialExceptionResolveView(APIView):
    """Resolve or reopen one serial exception with an audit note."""

    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        data = request.data
        try:
            record_id = int(data.get('id'))
        except (TypeError, ValueError):
            raise APIException({'detail': 'Serial record id is required'})
        action = str(data.get('action') or '').strip().upper()
        if action not in SERIAL_EXCEPTION_ACTIONS:
            raise APIException({
                'detail': 'Action must be ACCEPT_FOR_PUTAWAY, HOLD_QUARANTINE, REPAIR_REWORK, REJECT_RETURN, WAIVE_MISSING, or REOPEN'
            })
        record = AsnSerialRecord.objects.filter(id=record_id, openid=openid).first()
        if not record:
            raise APIException({'detail': 'Serial record does not exist'})
        is_missing = record.is_expected and not record.is_received
        is_exception = record.status in EXCEPTION_STATUSES
        if action != 'REOPEN' and not is_missing and not is_exception:
            raise APIException({'detail': 'This serial record has no open exception'})
        if action == 'WAIVE_MISSING' and not is_missing:
            raise APIException({'detail': 'WAIVE_MISSING is only valid for an expected SN that was not received'})
        if is_missing and action not in {'WAIVE_MISSING', 'REOPEN'}:
            raise APIException({
                'detail': 'A missing expected SN cannot be accepted or moved to putaway. Use WAIVE_MISSING only after the shortage is approved.',
                'code': 'MISSING_SN_NOT_PUTAWAY_ELIGIBLE',
            })
        if action != 'REOPEN' and not _resolution_note(data):
            raise APIException({'detail': 'A resolution note is required'})
        resolution_location = _clean(data.get('resolution_location'))
        if action in NON_PUTAWAY_RESOLUTIONS and not resolution_location:
            raise APIException({'detail': 'A hold or return location is required'})
        command, replay = consume_preview(
            request,
            'serial.resolve',
            request_payload(request),
            resource_id=str(record_id),
            asn_code=record.asn_code,
        )
        if replay is not None:
            return Response(replay)
        if action == 'REOPEN':
            if not record.exception_resolved:
                raise APIException({'detail': 'This serial exception is already open'})
            if record.exception_moved:
                raise APIException({'detail': 'A physically moved exception cannot be reopened; create a new reinspection record'})
            record.exception_resolved = False
            record.exception_resolution_action = ''
            record.exception_resolution_note = ''
            record.exception_resolution_location = ''
            record.exception_resolved_by = ''
            record.exception_resolved_at = None
        else:
            record.exception_resolved = True
            record.exception_resolution_action = action
            record.exception_resolution_note = _resolution_note(data)
            record.exception_resolution_location = resolution_location
            record.exception_resolved_by = _operator_name(request, openid)
            record.exception_resolved_at = timezone.now()
        record.save(update_fields=[
            'exception_resolved',
            'exception_resolution_action',
            'exception_resolution_note',
            'exception_resolution_location',
            'exception_resolved_by',
            'exception_resolved_at',
            'update_time',
        ])
        result = {
            'detail': 'Serial exception updated',
            'record': _record_json(record),
            'summary': _summary(openid, record.asn_code),
        }
        complete_preview(command, result)
        return Response(result)


EXCEPTION_BIN_PROPERTIES = {
    HOLD_QUARANTINE: {'holding', 'inspection'},
    REPAIR_REWORK: {'holding', 'inspection'},
    REJECT_RETURN: {'holding', 'damage'},
}


def _exception_bin(openid, action, requested_bin):
    bin_name = _clean(requested_bin)
    if not bin_name:
        raise APIException({'detail': 'A destination exception bin is required'})
    bin_detail = Bin.objects.filter(openid=openid, bin_name=bin_name, is_delete=False).first()
    if bin_detail is None:
        raise APIException({'detail': 'Exception destination bin does not exist'})
    if str(bin_detail.location_role or '').upper() == 'STAGING':
        raise APIException({'detail': 'Staging bins cannot receive exception inventory'})
    allowed = EXCEPTION_BIN_PROPERTIES.get(action, set())
    if str(bin_detail.bin_property or '').strip().lower() not in allowed:
        raise APIException({'detail': 'Bin property %s is not valid for %s' % (bin_detail.bin_property, action)})
    return bin_detail


def _move_exception_stock(openid, detail, quantity, bin_detail):
    assert_legacy_asn_putaway_allowed(openid, detail.asn_code, detail.goods_code)
    stock = StockListModel.objects.select_for_update().filter(
        openid=openid,
        goods_code=detail.goods_code,
    ).first()
    if stock is None:
        raise APIException({'detail': 'Stock record does not exist for %s' % detail.goods_code})
    if int(stock.sorted_stock or 0) < quantity:
        raise APIException({'detail': 'Not enough receiving-stage stock remains to move this exception'})
    remaining = int(detail.goods_actual_qty or 0) - int(detail.sorted_qty or 0)
    if remaining < quantity:
        raise APIException({'detail': 'Exception move quantity exceeds the remaining received quantity'})
    stock.sorted_stock = int(stock.sorted_stock or 0) - quantity
    stock.onhand_stock = int(stock.onhand_stock or 0) + quantity
    property_name = str(bin_detail.bin_property or '').strip().lower()
    if property_name == 'damage':
        stock.damage_stock = int(stock.damage_stock or 0) + quantity
    elif property_name == 'inspection':
        stock.inspect_stock = int(stock.inspect_stock or 0) + quantity
    else:
        stock.hold_stock = int(stock.hold_stock or 0) + quantity
    stock.save(update_fields=[
        'sorted_stock', 'onhand_stock', 'damage_stock', 'inspect_stock', 'hold_stock', 'update_time',
    ])
    StockBinModel.objects.create(
        openid=openid,
        bin_name=bin_detail.bin_name,
        goods_code=detail.goods_code,
        goods_desc=detail.goods_desc,
        goods_qty=quantity,
        bin_size=bin_detail.bin_size,
        bin_property=bin_detail.bin_property,
        t_code=Md5.md5('%s:%s' % (detail.goods_code, bin_detail.bin_name)),
        create_time=detail.create_time,
    )
    detail.sorted_qty = int(detail.sorted_qty or 0) + quantity
    if int(detail.sorted_qty or 0) >= int(detail.goods_actual_qty or 0):
        detail.asn_status = 5
    detail.save(update_fields=['sorted_qty', 'asn_status', 'update_time'])
    asn = AsnListModel.objects.select_for_update().filter(
        openid=openid,
        asn_code=detail.asn_code,
        is_delete=False,
    ).first()
    if asn and not AsnDetailModel.objects.filter(
        openid=openid,
        asn_code=detail.asn_code,
        asn_status=4,
        is_delete=False,
    ).exists():
        asn.asn_status = 5
        asn.save(update_fields=['asn_status', 'update_time'])
        from staging.models import StagingAssignment
        from staging.services import release_staging_slot
        release_staging_slot(openid, StagingAssignment.INBOUND, detail.asn_code)


class SerialExceptionMoveView(APIView):
    """Move one resolved serial exception out of receiving staging."""

    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        try:
            record_id = int(request.data.get('id'))
        except (TypeError, ValueError):
            raise APIException({'detail': 'Serial record id is required'})
        record = AsnSerialRecord.objects.select_for_update().filter(id=record_id, openid=openid).first()
        if record is None:
            raise APIException({'detail': 'Serial record does not exist'})
        if not record.exception_resolved or record.exception_resolution_action not in NON_PUTAWAY_RESOLUTIONS:
            raise APIException({'detail': 'Resolve the serial exception as HOLD, REPAIR, or REJECT before moving it'})
        if record.exception_moved:
            raise APIException({'detail': 'This serial exception has already been moved'})
        if not record.is_received:
            raise APIException({'detail': 'A missing serial has no physical unit to move'})
        bin_detail = _exception_bin(openid, record.exception_resolution_action, request.data.get('bin_name'))
        command, replay = consume_preview(
            request,
            'serial.exception_move',
            request_payload(request),
            resource_id=str(record_id),
            asn_code=record.asn_code,
        )
        if replay is not None:
            return Response(replay)
        detail = AsnDetailModel.objects.select_for_update().filter(
            openid=openid,
            asn_code=record.asn_code,
            goods_code=record.goods_code,
            is_delete=False,
        ).first()
        if detail is None:
            raise APIException({'detail': 'ASN detail does not exist for this serial'})
        _move_exception_stock(openid, detail, 1, bin_detail)
        record.exception_moved = True
        record.exception_move_bin = bin_detail.bin_name
        record.exception_moved_at = timezone.now()
        record.save(update_fields=['exception_moved', 'exception_move_bin', 'exception_moved_at', 'update_time'])
        result = {
            'detail': 'Serial exception moved out of staging',
            'record': _record_json(record),
            'destination_bin': bin_detail.bin_name,
            'summary': _summary(openid, record.asn_code),
        }
        complete_preview(command, result)
        return Response(result)


class QuantityExceptionMoveView(APIView):
    """Move a physical quantity exception to an explicit exception bin."""

    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        asn_code = _clean(request.data.get('asn_code'))
        goods_code = _clean(request.data.get('goods_code'))
        action = str(request.data.get('action') or '').strip().upper()
        try:
            quantity = int(request.data.get('qty'))
        except (TypeError, ValueError):
            raise APIException({'detail': 'qty must be a positive integer'})
        if action not in NON_PUTAWAY_RESOLUTIONS:
            raise APIException({'detail': 'Only HOLD_QUARANTINE, REPAIR_REWORK, or REJECT_RETURN can be physically moved'})
        if quantity <= 0 or not asn_code or not goods_code:
            raise APIException({'detail': 'ASN code, goods code, and positive qty are required'})
        detail = AsnDetailModel.objects.select_for_update().filter(
            openid=openid,
            asn_code=asn_code,
            goods_code=goods_code,
            is_delete=False,
        ).first()
        if detail is None or not detail.exception_resolved or detail.exception_resolution_action != action:
            raise APIException({'detail': 'Resolve the quantity exception with the same action before moving it'})
        bin_detail = _exception_bin(openid, action, request.data.get('bin_name'))
        command, replay = consume_preview(
            request,
            'serial.exception_move_quantity',
            request_payload(request),
            resource_id='%s:%s' % (asn_code, goods_code),
            asn_code=asn_code,
        )
        if replay is not None:
            return Response(replay)
        _move_exception_stock(openid, detail, quantity, bin_detail)
        movement = ExceptionQuantityMovement.objects.create(
            openid=openid,
            asn_code=asn_code,
            goods_code=goods_code,
            quantity=quantity,
            action=action,
            bin_name=bin_detail.bin_name,
            operator=_operator_name(request, openid),
        )
        result = {
            'detail': 'Quantity exception moved out of staging',
            'movement_id': movement.id,
            'asn_code': asn_code,
            'goods_code': goods_code,
            'qty': quantity,
            'destination_bin': bin_detail.bin_name,
            'summary': _summary(openid, asn_code),
        }
        complete_preview(command, result)
        return Response(result)


class QuantityExceptionResolveView(APIView):
    """Resolve or reopen the quantity variance recorded during QC."""

    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        data = request.data
        asn_code = _clean(data.get('asn_code'))
        goods_code = _clean(data.get('goods_code'))
        action = str(data.get('action') or '').strip().upper()
        if not asn_code or not goods_code:
            raise APIException({'detail': 'ASN Code and Goods Code are required'})
        if action not in {
            ACCEPT_FOR_PUTAWAY,
            LEGACY_ACCEPT_EXCEPTION,
            HOLD_QUARANTINE,
            REPAIR_REWORK,
            REJECT_RETURN,
            'REOPEN',
        }:
            raise APIException({
                'detail': 'Action must be ACCEPT_FOR_PUTAWAY, HOLD_QUARANTINE, REPAIR_REWORK, REJECT_RETURN, or REOPEN'
            })
        detail = AsnDetailModel.objects.filter(
            openid=openid,
            asn_code=asn_code,
            goods_code=goods_code,
            is_delete=False,
        ).first()
        if not detail:
            raise APIException({'detail': 'ASN detail does not exist'})
        quantity = int(detail.goods_shortage_qty or 0) + int(detail.goods_more_qty or 0) + int(detail.goods_damage_qty or 0)
        if action != 'REOPEN' and quantity <= 0:
            raise APIException({'detail': 'This ASN detail has no quantity exception'})
        if action != 'REOPEN' and not _resolution_note(data):
            raise APIException({'detail': 'A resolution note is required'})
        resolution_location = _clean(data.get('resolution_location'))
        if action in NON_PUTAWAY_RESOLUTIONS and not resolution_location:
            raise APIException({'detail': 'A hold or return location is required'})
        command, replay = consume_preview(
            request,
            'serial.resolve_quantity',
            request_payload(request),
            resource_id='%s:%s' % (asn_code, goods_code),
            asn_code=asn_code,
        )
        if replay is not None:
            return Response(replay)
        if action == 'REOPEN':
            if not detail.exception_resolved:
                raise APIException({'detail': 'This quantity exception is already open'})
            detail.exception_resolved = False
            detail.exception_resolution_action = ''
            detail.exception_resolution_note = ''
            detail.exception_resolution_location = ''
            detail.exception_resolved_by = ''
            detail.exception_resolved_at = None
        else:
            detail.exception_resolved = True
            detail.exception_resolution_action = action
            detail.exception_resolution_note = _resolution_note(data)
            detail.exception_resolution_location = resolution_location
            detail.exception_resolved_by = _operator_name(request, openid)
            detail.exception_resolved_at = timezone.now()
        detail.save(update_fields=[
            'exception_resolved',
            'exception_resolution_action',
            'exception_resolution_note',
            'exception_resolution_location',
            'exception_resolved_by',
            'exception_resolved_at',
            'update_time',
        ])
        result = {
            'detail': 'Quantity exception updated',
            'asn_detail_id': detail.id,
            'asn_code': detail.asn_code,
            'goods_code': detail.goods_code,
            'exception_resolved': detail.exception_resolved,
            'exception_resolution_action': detail.exception_resolution_action,
            'exception_resolution_note': detail.exception_resolution_note,
            'resolution_location': detail.exception_resolution_location,
            'summary': _summary(openid, detail.asn_code),
        }
        complete_preview(command, result)
        return Response(result)


class SerialSummaryView(APIView):
    def get(self, request):
        openid = _openid(request)
        asn_code = _clean(request.query_params.get('asn_code'))
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        _asn = AsnListModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False).first()
        if not _asn:
            raise APIException({'detail': 'ASN Code does not exists'})
        return Response(_summary(openid, asn_code))


class ExpectedSerialView(APIView):
    def post(self, request):
        openid = _openid(request)
        data = request.data
        asn_code = _clean(data.get('asn_code'))
        default_goods = _clean(data.get('goods_code'))
        rows = data.get('rows') or []
        if not rows:
            rows = [{'serial_number': value, 'goods_code': default_goods} for value in (data.get('serial_numbers') or [])]
        if not asn_code or not rows:
            raise APIException({'detail': 'ASN Code and serial rows are required'})
        created = 0
        updated = 0
        results = []
        with transaction.atomic():
            for row in rows:
                if isinstance(row, str):
                    row = {'serial_number': row, 'goods_code': default_goods}
                goods_code = _clean(row.get('goods_code') or default_goods)
                record, was_created = _save_expected(openid, request, asn_code, goods_code, row.get('serial_number'), row=row)
                created += int(was_created)
                updated += int(not was_created)
                results.append(_record_json(record))
        return Response({'detail': 'success', 'created': created, 'updated': updated, 'results': results, 'summary': _summary(openid, asn_code)})


class ScanSerialView(APIView):
    def post(self, request):
        openid = _openid(request)
        data = request.data
        asn_code = _clean(data.get('asn_code'))
        goods_code = _clean(data.get('goods_code'))
        record, created = _scan(
            openid,
            request,
            asn_code,
            goods_code,
            data.get('serial_number'),
            damaged=bool(data.get('damaged', False)),
            row=data,
        )
        return Response({'detail': 'success', 'created': created, 'record': _record_json(record), 'summary': _summary(openid, asn_code)})


def _header_key(value):
    return ''.join(char for char in str(value or '').upper() if char.isalnum())


def _first_column(index, names):
    for name in names:
        if _header_key(name) in index:
            return index[_header_key(name)]
    return None


def _serial_rows_from_workbook(file_bytes, inbound_po='', shipout_ref=''):
    """Read every SKU/SN table section from the customer acceptance workbook."""
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise APIException({'detail': 'Unable to read Excel file: ' + str(exc)})

    rows = []
    try:
        for sheet in workbook.worksheets:
            section_index = None
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                headers = {
                    _header_key(value): position
                    for position, value in enumerate(values)
                    if _header_key(value)
                }
                if (
                    _first_column(headers, ('SKU#', 'SKU', 'Part Number', 'Goods Code', 'Item')) is not None
                    and _first_column(headers, ('SN#', 'SN', 'Serial Number', 'Serial', 'Serial No')) is not None
                ):
                    section_index = headers
                    continue
                if section_index is None or not any(value not in (None, '') for value in values):
                    continue

                def value(*names):
                    pos = _first_column(section_index, names)
                    return values[pos] if pos is not None and pos < len(values) else ''

                row_po = _clean(value('Inbound PO#', 'Inbound PO', 'PO#'))
                row_shipout = _clean(value('SHIPOUT#', 'Shipout Ref', 'Shipout'))
                if inbound_po and row_po != inbound_po:
                    continue
                if shipout_ref and row_shipout != shipout_ref:
                    continue
                if _first_column(section_index, ('Inbound PO#', 'Inbound PO', 'PO#')) is not None and not row_po and not shipout_ref:
                    continue
                goods_code = _clean(value('SKU#', 'SKU', 'Part Number', 'Goods Code', 'Item'))
                serial_number = _clean(value('SN#', 'SN', 'Serial Number', 'Serial', 'Serial No'))
                if not goods_code or not serial_number:
                    continue
                rows.append({
                    'sheet': sheet.title,
                    'row_number': row_number,
                    'values': values,
                    'index': section_index,
                })
    finally:
        workbook.close()

    return rows


def _pack_list_rows_from_workbook(upload):
    try:
        file_bytes = upload.read()
        upload.seek(0)
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise APIException({'detail': 'Unable to read Pack List Excel file: ' + str(exc)})
    if not values:
        raise APIException({'detail': 'Pack List Excel file is empty'})
    index = {
        _header_key(value): position
        for position, value in enumerate(values[0])
        if _header_key(value)
    }
    sku_column = _first_column(index, ('SKU#', 'SKU', 'Goods Code', 'GoodsCode', 'Item', 'Part Number'))
    qty_column = _first_column(index, ('Item Qty', 'Qty', 'Quantity', 'Goods Qty', 'ASN Qty', 'Total Qty'))
    serial_column = _first_column(index, ('SN#', 'SN', 'Serial Number', 'Serial', 'Serial No'))
    customer_sku_column = _first_column(index, ('Customer SKU', 'Customer Part Number', 'Customer Item'))
    customer_ssku_column = _first_column(index, ('S-SKU', 'Customer S-SKU', 'Client SKU'))
    package_type_column = _first_column(index, ('Package Type', 'Package Code', 'Package ID', 'Package'))
    desc_column = _first_column(index, ('Description', 'Goods Description', 'Product Description'))
    weight_column = _first_column(index, ('Weight', 'Goods Weight'))
    volume_column = _first_column(index, ('Volume', 'Goods Volume'))
    total_column = _first_column(index, ('Total', 'Total Qty', 'Item Total'))
    if sku_column is None:
        raise APIException({'detail': 'Pack List must contain a SKU or Goods Code column'})
    rows = []
    for row_number, values_row in enumerate(values[1:], start=2):
        def value_at(column):
            return values_row[column] if column is not None and column < len(values_row) else ''

        goods_code = _clean(value_at(sku_column))
        serial_number = _clean(value_at(serial_column))
        if not goods_code and not serial_number:
            continue
        qty_value = value_at(qty_column)
        qty = int(_number(qty_value, Decimal('1') if serial_number else Decimal('0')))
        rows.append({
            'goods_code': goods_code,
            'customer_goods_code': _clean(value_at(customer_sku_column)),
            'customer_ssku': _clean(value_at(customer_ssku_column)),
            'package_type': _clean(value_at(package_type_column)),
            'serial_number': serial_number,
            'goods_qty': qty,
            'total_qty': int(_number(value_at(total_column), Decimal(str(qty))) or qty),
            'goods_desc': str(value_at(desc_column) or '').strip(),
            'goods_weight': _number(value_at(weight_column)),
            'goods_volume': _number(value_at(volume_column)),
            'source_row': row_number,
        })
    if not rows:
        raise APIException({'detail': 'Pack List contains no usable rows'})
    return rows, sha256(file_bytes).hexdigest()


def _validate_pack_list_rows(openid, asn_code, rows):
    asn = AsnListModel.objects.filter(openid=openid, asn_code=asn_code, is_delete=False).first()
    if not asn:
        raise APIException({'detail': 'ASN Code does not exists'})
    normalized_rows = []
    quantities = {}
    serial_numbers = set()
    has_serials = False
    for position, raw_row in enumerate(rows, start=1):
        if isinstance(raw_row, str):
            raw_row = {'goods_code': raw_row, 'goods_qty': 1}
        goods_code = _clean(raw_row.get('goods_code'))
        serial_number = _clean(raw_row.get('serial_number'))
        if not goods_code:
            raise APIException({'detail': 'Pack List row %s is missing internal SKU' % position})
        _, detail = _asn_detail(openid, asn_code, goods_code)
        qty = int(_number(raw_row.get('goods_qty'), Decimal('1') if serial_number else Decimal('0')))
        if qty <= 0:
            raise APIException({'detail': 'Pack List row %s quantity must be greater than 0' % position})
        quantities[goods_code] = quantities.get(goods_code, 0) + qty
        if quantities[goods_code] > int(detail.goods_qty):
            raise APIException({'detail': 'Pack List quantity exceeds ASN quantity for SKU ' + goods_code})
        has_serials = has_serials or bool(serial_number)
        if serial_number:
            if serial_number in serial_numbers:
                raise APIException({'detail': 'Pack List contains duplicate Serial Number ' + serial_number})
            serial_numbers.add(serial_number)
        normalized_rows.append({
            'goods_code': goods_code,
            'customer_goods_code': _clean(raw_row.get('customer_goods_code')),
            'customer_ssku': _clean(raw_row.get('customer_ssku')),
            'package_type': _clean(raw_row.get('package_type')),
            'serial_number': serial_number,
            'goods_qty': qty,
            'total_qty': int(_number(raw_row.get('total_qty'), Decimal(str(qty))) or qty),
            'goods_desc': str(raw_row.get('goods_desc') or '').strip(),
            'goods_weight': _number(raw_row.get('goods_weight')),
            'goods_volume': _number(raw_row.get('goods_volume')),
            'source_row': int(raw_row.get('source_row') or position),
        })
    return {
        'asn': asn,
        'rows': normalized_rows,
        'has_serials': has_serials,
        'total_qty': sum(row['goods_qty'] for row in normalized_rows),
        'expected_serial_count': len(serial_numbers),
    }


def _pack_list_preview_json(asn_code, validation, package_qty, content_hash, duplicate_document=None, current_document=None, receiving_started=False):
    return {
        'asn_code': asn_code,
        'status': 'DUPLICATE' if duplicate_document else 'PREVIEW',
        'content_hash': content_hash,
        'row_count': len(validation['rows']),
        'total_qty': validation['total_qty'],
        'has_serials': validation['has_serials'],
        'expected_serial_count': validation['expected_serial_count'],
        'package_qty': max(0, int(package_qty or 0)),
        'duplicate_document': _pack_list_json(duplicate_document) if duplicate_document else None,
        'current_document': _pack_list_json(current_document) if current_document else None,
        'replace_required': bool(current_document and not duplicate_document),
        'receiving_started': receiving_started,
        'late_reference_required': bool(receiving_started and not duplicate_document),
        'lines': [
            {
                'goods_code': row['goods_code'],
                'customer_goods_code': row['customer_goods_code'],
                'customer_ssku': row['customer_ssku'],
                'package_type': row['package_type'],
                'goods_qty': row['goods_qty'],
                'total_qty': row['total_qty'],
                'serial_number': row['serial_number'],
                'goods_desc': row['goods_desc'],
                'source_row': row['source_row'],
            }
            for row in validation['rows']
        ],
    }


def _create_pack_list(openid, request, asn_code, rows, source_type='AI_AGENT', content_hash='', note='', package_qty=0, replace=False, late_reference=False):
    validation = _validate_pack_list_rows(openid, asn_code, rows)
    asn = validation['asn']
    normalized_rows = validation['rows']
    has_serials = validation['has_serials']
    package_qty = max(0, int(package_qty or 0))
    source_type = source_type if source_type in dict(PackListDocument.SOURCE_TYPES) else 'MANUAL'
    document = PackListDocument.objects.select_for_update().filter(
        openid=openid,
        asn_code=asn_code,
        is_current=True,
    ).first()
    receiving_started = _receiving_started(openid, asn_code)
    next_version = 1
    if document:
        if document.content_hash == str(content_hash or ''):
            return document, None, False
        if not replace:
            raise APIException({
                'detail': 'A different Pack List already exists for this ASN. Preview the differences and use the explicit Replace action.',
                'code': 'PACK_LIST_REPLACE_REQUIRED',
                'document_id': document.id,
            })
        if receiving_started and not late_reference:
            raise APIException({
                'detail': 'Receiving has started; import this Pack List as a late reference revision.',
                'code': 'PACK_LIST_LATE_REFERENCE_REQUIRED',
            })
        if receiving_started:
            next_version = int(document.version or 0) + 1
            document.is_current = False
            document.status = PackListDocument.ARCHIVED
            document.save(update_fields=['is_current', 'status', 'update_time'])
            document = None
            late_reference = True
        else:
            document.serial_records.filter(is_expected=True, is_received=False).update(
                pack_list=None,
                is_expected=False,
                expected_goods_code='',
                status=AsnSerialRecord.UNVERIFIED,
            )
            document.lines.filter(is_current=True).update(is_current=False)
            document.version = int(document.version or 0) + 1
            document.has_serials = has_serials
            document.package_qty = package_qty
            document.note = str(note or '')
            document.source_type = source_type
            document.content_hash = str(content_hash or '')[:64]
            document.confirmed_by = ''
            document.confirmed_at = None
            document.status = PackListDocument.PENDING
            document.late_reference = False
            document.save(update_fields=[
                'version', 'source_type', 'content_hash', 'status', 'has_serials',
                'package_qty', 'note', 'late_reference', 'confirmed_by', 'confirmed_at', 'update_time',
            ])
    created_document = document is None
    if created_document:
        document = PackListDocument.objects.create(
            openid=openid,
            asn_code=asn_code,
            version=next_version,
            source_type=source_type,
            content_hash=str(content_hash or '')[:64],
            is_current=True,
            has_serials=has_serials,
            package_qty=package_qty,
            note=str(note or ''),
            late_reference=bool(late_reference or receiving_started),
            created_by=_operator_name(request, openid),
        )
    import_batch = None
    if content_hash:
        import_batch = PackListImportBatch.objects.create(
            openid=openid,
            asn_code=asn_code,
            import_type=PackListImportBatch.PACK_LIST,
            content_hash=str(content_hash)[:64],
            row_count=len(normalized_rows),
            source_type=source_type,
            imported_by=_operator_name(request, openid),
            note=str(note or ''),
        )
        document.import_batch = import_batch
        document.save(update_fields=['import_batch', 'update_time'])
    if document.package_qty > 0 and int(asn.package_qty or 0) != document.package_qty:
        asn.package_qty = document.package_qty
        asn.save(update_fields=['package_qty', 'update_time'])
    for row in normalized_rows:
        PackListLine.objects.create(
            pack_list=document,
            openid=openid,
            asn_code=asn_code,
            goods_code=row['goods_code'],
            customer_goods_code=row['customer_goods_code'],
            is_current=True,
            customer_ssku=row['customer_ssku'],
            goods_qty=row['goods_qty'],
            total_qty=row['total_qty'],
            package_type=row['package_type'],
            goods_desc=row['goods_desc'],
            goods_weight=row['goods_weight'],
            goods_volume=row['goods_volume'],
            source_row=row['source_row'],
        )
        if row['serial_number']:
            _save_expected(
                openid,
                request,
                asn_code,
                row['goods_code'],
                row['serial_number'],
                row=row,
                source='pack_list',
                pack_list=document,
                import_batch=import_batch,
            )
    return document, import_batch, created_document


def _reconcile_pack_list(document):
    expected = {
        record.serial_number: record.goods_code
        for record in document.serial_records.filter(is_expected=True)
    }
    records = AsnSerialRecord.objects.filter(openid=document.openid, asn_code=document.asn_code)
    for record in records.filter(status__in=[AsnSerialRecord.UNVERIFIED, AsnSerialRecord.UNEXPECTED]):
        expected_goods_code = expected.get(record.serial_number)
        record.pack_list = document
        if expected_goods_code and expected_goods_code == record.goods_code:
            record.is_expected = True
            record.expected_goods_code = expected_goods_code
            if record.damaged:
                record.status = AsnSerialRecord.DAMAGED
            elif record.scan_count > 1:
                record.status = AsnSerialRecord.DUPLICATE
            else:
                record.status = AsnSerialRecord.ACCEPTED
        elif document.has_serials:
            record.status = AsnSerialRecord.UNEXPECTED
        record.save(update_fields=['pack_list', 'is_expected', 'expected_goods_code', 'status', 'update_time'])


class PackListListView(APIView):
    def get(self, request):
        openid = _openid(request)
        asn_code = _clean(request.query_params.get('asn_code'))
        documents = PackListDocument.objects.filter(openid=openid)
        if asn_code:
            documents = documents.filter(asn_code=asn_code)
        return Response({
            'count': documents.count(),
            'results': [_pack_list_json(document) for document in documents],
            'summary': _summary(openid, asn_code) if asn_code else None,
            'inspection_batches': [
                _inspection_batch_json(batch)
                for batch in PackListImportBatch.objects.filter(
                    openid=openid,
                    asn_code=asn_code,
                    import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
                )[:50]
            ] if asn_code else [],
        })


class PackListCreateView(APIView):
    def post(self, request):
        openid = _openid(request)
        data = request.data
        asn_code = _clean(data.get('asn_code'))
        rows = data.get('rows') or []
        if not asn_code or not rows:
            raise APIException({'detail': 'ASN Code and Pack List rows are required'})
        with transaction.atomic():
            document, _, _ = _create_pack_list(
                openid,
                request,
                asn_code,
                rows,
                source_type=str(data.get('source_type') or 'AI_AGENT').upper(),
                note=data.get('note'),
                package_qty=data.get('package_qty'),
                replace=str(data.get('replace', '')).lower() == 'true',
                late_reference=str(data.get('late_reference', '')).lower() == 'true',
            )
        return Response({'detail': 'success', 'document': _pack_list_json(document), 'summary': _summary(openid, asn_code)})


class PackListPreviewView(APIView):
    def post(self, request):
        openid = _openid(request)
        upload = request.FILES.get('file')
        asn_code = _clean(request.data.get('asn_code'))
        if not upload:
            raise ValidationError({'detail': 'Pack List Excel file is required'})
        if upload.size > 20 * 1024 * 1024:
            raise ValidationError({'detail': 'Pack List file is too large'})
        if not asn_code:
            raise ValidationError({'detail': 'ASN Code is required'})
        try:
            rows, content_hash = _pack_list_rows_from_workbook(upload)
            validation = _validate_pack_list_rows(openid, asn_code, rows)
        except APIException as exc:
            raise ValidationError(exc.detail)
        current_document = PackListDocument.objects.filter(
            openid=openid,
            asn_code=asn_code,
            is_current=True,
        ).first()
        duplicate_document = current_document if current_document and current_document.content_hash == content_hash else None
        receiving_started = _receiving_started(openid, asn_code)
        result = {
            'detail': 'preview',
            'preview': _pack_list_preview_json(
                asn_code,
                validation,
                request.data.get('package_qty'),
                content_hash,
                duplicate_document=duplicate_document,
                current_document=current_document,
                receiving_started=receiving_started,
            ),
        }
        if is_agent_request(request):
            result['agent'] = create_preview(
                request,
                'packlist.import',
                {
                    'asn_code': asn_code,
                    'content_hash': content_hash,
                    'package_qty': str(request.data.get('package_qty') or ''),
                    'source_type': str(request.data.get('source_type') or 'AI_AGENT').upper(),
                    'note': str(request.data.get('note') or ''),
                    'replace': str(request.data.get('replace', '')).lower() == 'true',
                    'late_reference': str(request.data.get('late_reference', '')).lower() == 'true',
                },
                asn_code=asn_code,
            )
        return Response(result)


class PackListImportView(APIView):
    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        upload = request.FILES.get('file')
        asn_code = _clean(request.data.get('asn_code'))
        if not upload:
            raise APIException({'detail': 'Pack List Excel file is required'})
        if upload.size > 20 * 1024 * 1024:
            raise APIException({'detail': 'Pack List file is too large'})
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        rows, content_hash = _pack_list_rows_from_workbook(upload)
        validation = _validate_pack_list_rows(openid, asn_code, rows)
        command, replay = consume_preview(
            request,
            'packlist.import',
            {
                'asn_code': asn_code,
                'content_hash': content_hash,
                'package_qty': str(request.data.get('package_qty') or ''),
                'source_type': str(request.data.get('source_type') or 'AI_AGENT').upper(),
                'note': str(request.data.get('note') or ''),
                'replace': str(request.data.get('replace', '')).lower() == 'true',
                'late_reference': str(request.data.get('late_reference', '')).lower() == 'true',
            },
            asn_code=asn_code,
        )
        if replay is not None:
            return Response(replay)
        existing_document = PackListDocument.objects.filter(
            openid=openid,
            asn_code=asn_code,
            is_current=True,
        ).first()
        if existing_document and existing_document.content_hash == content_hash:
            result = {
                'detail': 'already_exists',
                'duplicate': True,
                'document': _pack_list_json(existing_document),
                'summary': _summary(openid, asn_code),
            }
            complete_preview(command, result)
            return Response(result)
        replaced = bool(existing_document)
        receiving_started = _receiving_started(openid, asn_code)
        late_reference = str(request.data.get('late_reference', '')).lower() == 'true'
        with transaction.atomic():
            document, _, created = _create_pack_list(
                openid,
                request,
                asn_code,
                rows,
                source_type=str(request.data.get('source_type') or 'AI_AGENT').upper(),
                content_hash=content_hash,
                note=request.data.get('note'),
                package_qty=request.data.get('package_qty'),
                replace=str(request.data.get('replace', '')).lower() == 'true',
                late_reference=late_reference,
            )
        result = {
            'detail': 'success' if created else 'already_exists',
            'duplicate': not created,
            'replaced': replaced,
            'late_reference': late_reference or receiving_started,
            'document': _pack_list_json(document),
            'summary': _summary(openid, asn_code),
        }
        complete_preview(command, result)
        return Response(result)


class PackListConfirmView(APIView):
    @transaction.atomic
    def post(self, request):
        openid = _openid(request)
        try:
            document_id = int(request.data.get('id'))
        except (TypeError, ValueError):
            raise APIException({'detail': 'Pack List id is required'})
        document = PackListDocument.objects.filter(id=document_id, openid=openid, is_current=True).first()
        if not document:
            raise APIException({'detail': 'Pack List does not exist'})
        command, replay = consume_preview(
            request,
            'packlist.confirm',
            request_payload(request),
            resource_id=str(document_id),
            asn_code=document.asn_code,
        )
        if replay is not None:
            return Response(replay)
        with transaction.atomic():
            document.status = PackListDocument.CONFIRMED
            document.confirmed_by = _operator_name(request, openid)
            document.confirmed_at = timezone.now()
            document.save(update_fields=['status', 'confirmed_by', 'confirmed_at', 'update_time'])
            _reconcile_pack_list(document)
        result = {'detail': 'success', 'document': _pack_list_json(document), 'summary': _summary(openid, document.asn_code)}
        complete_preview(command, result)
        return Response(result)


class SerialImportPreviewView(APIView):
    """Parse an acceptance workbook without writing SN/QC or inventory data."""

    def post(self, request, inspection=False):
        openid = _openid(request)
        upload = request.FILES.get('file')
        asn_code = _clean(request.data.get('asn_code'))
        if not upload:
            raise ValidationError({'detail': 'Excel file is required'})
        if upload.size > 10 * 1024 * 1024:
            raise ValidationError({'detail': 'Excel file is too large'})
        mode = 'receive' if inspection else str(request.data.get('mode') or 'expected').lower()
        inbound_po = _clean(request.data.get('inbound_po'))
        shipout_ref = _clean(request.data.get('shipout_ref'))
        allow_all = str(request.data.get('allow_all', '')).lower() == 'true'
        if not asn_code:
            raise ValidationError({'detail': 'ASN Code is required'})
        if not inbound_po and not shipout_ref and not allow_all:
            raise ValidationError({'detail': 'Provide inbound_po or shipout_ref before importing a mixed scan sheet'})
        try:
            file_bytes = upload.read()
            candidate_rows = _serial_rows_from_workbook(file_bytes, inbound_po, shipout_ref)
        except APIException as exc:
            raise ValidationError(exc.detail)
        except Exception as exc:
            raise ValidationError({'detail': 'Unable to read Excel file: ' + str(exc)})
        if not candidate_rows:
            raise ValidationError({
                'detail': 'No matching SKU/SN rows were found in the acceptance workbook; import was not created',
                'code': 'QC_IMPORT_NO_MATCH',
            })
        content_hash = sha256((mode + ':').encode('utf-8') + file_bytes).hexdigest()
        operation = 'inspection.import' if inspection else 'serial.import'
        payload = {
            'asn_code': asn_code,
            'mode': mode,
            'content_hash': content_hash,
            'inbound_po': inbound_po,
            'shipout_ref': shipout_ref,
            'allow_all': allow_all,
            'source_type': str(request.data.get('source_type') or 'AI_AGENT').upper(),
            'note': str(request.data.get('note') or ''),
            'evidence_url': _text(request.data.get('evidence_url')),
        }
        result = {
            'detail': 'preview',
            'operation': operation,
            'asn_code': asn_code,
            'mode': mode,
            'content_hash': content_hash,
            'matched_rows': len(candidate_rows),
            'sample': [
                {
                    'sheet': row.get('sheet'),
                    'row_number': row.get('row_number'),
                    'values': row.get('values'),
                }
                for row in candidate_rows[:20]
            ],
        }
        if is_agent_request(request):
            result['agent'] = create_preview(request, operation, payload, asn_code=asn_code)
        return Response(result)


class SerialImportView(APIView):
    @transaction.atomic
    def post(self, request, inspection=False):
        openid = _openid(request)
        upload = request.FILES.get('file')
        if not upload:
            raise APIException({'detail': 'Excel file is required'})
        if upload.size > 10 * 1024 * 1024:
            raise APIException({'detail': 'Excel file is too large'})
        mode = 'receive' if inspection else str(request.data.get('mode') or 'expected').lower()
        if mode not in ('expected', 'receive'):
            raise APIException({'detail': 'Mode must be expected or receive'})
        asn_code = _clean(request.data.get('asn_code'))
        inbound_po = _clean(request.data.get('inbound_po'))
        shipout_ref = _clean(request.data.get('shipout_ref'))
        evidence_url = _text(request.data.get('evidence_url'))
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        if not inbound_po and not shipout_ref and str(request.data.get('allow_all', '')).lower() != 'true':
            raise APIException({'detail': 'Provide inbound_po or shipout_ref before importing a mixed scan sheet'})
        try:
            file_bytes = upload.read()
            upload.seek(0)
            candidate_rows = _serial_rows_from_workbook(file_bytes, inbound_po, shipout_ref)
        except Exception as exc:
            if isinstance(exc, APIException):
                raise
            raise APIException({'detail': 'Unable to read Excel file: ' + str(exc)})
        if not candidate_rows:
            raise APIException({
                'detail': 'No matching SKU/SN rows were found in the acceptance workbook; import was not created',
                'code': 'QC_IMPORT_NO_MATCH',
            })
        content_hash = sha256((mode + ':' ).encode('utf-8') + file_bytes).hexdigest()
        asn = AsnListModel.objects.select_for_update().filter(
            openid=openid,
            asn_code=asn_code,
            is_delete=False,
        ).first()
        if asn is None:
            raise APIException({'detail': 'ASN Code does not exists'})
        operation = 'inspection.import' if inspection else 'serial.import'
        command, replay = consume_preview(
            request,
            operation,
            {
                'asn_code': asn_code,
                'mode': mode,
                'content_hash': content_hash,
                'inbound_po': inbound_po,
                'shipout_ref': shipout_ref,
                'allow_all': str(request.data.get('allow_all', '')).lower() == 'true',
                'source_type': str(request.data.get('source_type') or 'AI_AGENT').upper(),
                'note': str(request.data.get('note') or ''),
                'evidence_url': evidence_url,
            },
            asn_code=asn_code,
        )
        if replay is not None:
            return Response(replay)
        existing_batch = PackListImportBatch.objects.select_for_update().filter(
            openid=openid,
            asn_code=asn_code,
            import_type=(
                PackListImportBatch.RECEIVING_ACCEPTANCE
                if mode == 'receive' else PackListImportBatch.EXPECTED_SERIALS
            ),
            content_hash=content_hash,
        ).first()
        if existing_batch:
            result = {
                'detail': 'already_exists',
                'duplicate': True,
                'mode': mode,
                'batch_id': existing_batch.id,
                'matched_rows': existing_batch.row_count,
                'created': 0,
                'updated': 0,
                'skipped': 0,
                'errors': [],
                'batch': _inspection_batch_json(existing_batch),
                'summary': _summary(openid, asn_code),
            }
            complete_preview(command, result)
            return Response(result)
        import_batch = PackListImportBatch.objects.create(
            openid=openid,
            asn_code=asn_code,
            import_type=(
                PackListImportBatch.RECEIVING_ACCEPTANCE
                if mode == 'receive' else PackListImportBatch.EXPECTED_SERIALS
            ),
            content_hash=content_hash,
            imported_by=_operator_name(request, openid),
            note=str(request.data.get('note') or ('QC inspection import' if mode == 'receive' else 'Expected serial import')),
            evidence_url=evidence_url,
            source_type=str(request.data.get('source_type') or 'AI_AGENT').upper(),
        )
        matched = 0
        created = 0
        updated = 0
        skipped = 0
        errors = []
        for candidate in candidate_rows:
            row_number = candidate['row_number']
            values = candidate['values']
            index = candidate['index']

            def value(*names):
                pos = _first_column(index, names)
                return values[pos] if pos is not None and pos < len(values) else ''

            row_po = _clean(value('Inbound PO#', 'Inbound PO', 'PO#'))
            row_shipout = _clean(value('SHIPOUT#', 'Shipout Ref', 'Shipout'))
            goods_code = _clean(value('SKU#', 'SKU', 'Part Number', 'Goods Code', 'Item'))
            serial_number = _clean(value('SN#', 'SN', 'Serial Number', 'Serial', 'Serial No'))
            matched += 1
            inspection_result = value('Inspection Result', 'QC Result', 'Check Result', 'Condition', 'Result')
            row_data = {
                'double_scan_sn': value('Double-Scan SN#', 'Double Scan SN', 'Double-Scan SN'),
                'inbound_po': row_po,
                'inbound_date': value('Inbound Date', 'Date'),
                'source_location': value('Location'),
                'shipout_ref': row_shipout,
                'damaged': _is_damage_flag(value('Damaged', 'Damage', 'Damage Flag', 'Damage Status')) or _is_damage_flag(inspection_result),
                'note': value('QC Note', 'Inspection Note', 'Check Note', 'Note', 'Remarks') or inspection_result,
                'evidence_url': _text(value('Evidence URL', 'Photo URL', 'Video URL', 'Google Drive', 'Google Drive URL')) or evidence_url,
                'source_row': row_number,
                'import_batch': import_batch,
            }
            try:
                if mode == 'expected':
                    record, was_created = _save_expected(openid, request, asn_code, goods_code, serial_number, row=row_data, source='excel', import_batch=import_batch)
                else:
                    record, was_created = _scan(
                        openid,
                        request,
                        asn_code,
                        goods_code,
                        serial_number,
                        damaged=row_data['damaged'],
                        row=row_data,
                        source='inspection' if mode == 'receive' else 'excel',
                        import_batch=import_batch,
                    )
                created += int(was_created)
                updated += int(not was_created)
            except Exception as exc:
                skipped += 1
                if len(errors) < 50:
                    errors.append({'row': row_number, 'sku': goods_code, 'sn': serial_number, 'detail': str(exc)})
        import_batch.row_count = matched
        import_batch.matched_count = matched
        touched_records = AsnSerialRecord.objects.filter(import_batch=import_batch)
        import_batch.accepted_count = touched_records.filter(status=AsnSerialRecord.ACCEPTED).count()
        import_batch.exception_count = touched_records.filter(
            status__in=EXCEPTION_STATUSES,
            exception_resolved=False,
        ).count()
        import_batch.status = (
            PackListImportBatch.PARTIAL if errors else
            PackListImportBatch.EXCEPTION if import_batch.exception_count else
            PackListImportBatch.PASSED
        )
        import_batch.save(update_fields=[
            'row_count', 'matched_count', 'accepted_count', 'exception_count', 'status',
        ])
        result = {
            'detail': 'success' if not errors else 'partial_success',
            'mode': mode,
            'batch_id': import_batch.id,
            'matched_rows': matched,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors,
            'batch': _inspection_batch_json(import_batch),
            'summary': _summary(openid, asn_code),
        }
        complete_preview(command, result)
        return Response(result)


class InspectionBatchListView(APIView):
    """Return QC inspection import history without exposing uploaded files."""

    def get(self, request):
        openid = _openid(request)
        asn_code = _clean(request.query_params.get('asn_code'))
        if not asn_code:
            raise APIException({'detail': 'ASN Code is required'})
        batches = PackListImportBatch.objects.filter(
            openid=openid,
            asn_code=asn_code,
            import_type=PackListImportBatch.RECEIVING_ACCEPTANCE,
        )[:50]
        return Response({
            'count': batches.count(),
            'results': [_inspection_batch_json(batch) for batch in batches],
        })
