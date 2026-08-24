#!/usr/bin/env python3
"""Rebuild GreaterWMS input data from the locally exported session files.

This is intentionally an offline transformation. It does not connect to a
database or a WMS service; the generated folder is the recovery source for a
later, authenticated API import.
"""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "project" / "dallas peter warehouse" / "session-backup" / "2026-08-07-greaterwms-inputs"
MAIL = Path("/private/tmp/greaterwms-mail")
SKU_SOURCE = MAIL / "Export Standalone_Items.xlsx"
SKU_SOURCE_EMPTY = MAIL / "Export Standalone_Items (1).xlsx"
SKU_DELTA_SOURCE = MAIL / "latest_delta_sku.xlsx"
SCAN_SOURCE = MAIL / "Scan sheet sample.xlsx"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def serialise(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def convert_dimension(value: float, unit: str) -> float:
    if unit == "cm":
        return value / 2.54
    if unit == "mm":
        return value / 25.4
    return value


def convert_weight(value: float, unit: str) -> float:
    if unit == "kg":
        return value * 2.2046226218
    if unit == "g":
        return value / 453.59237
    if unit == "oz":
        return value / 16
    return value


def normalise_sku(row: dict[str, Any], source_row: int, source: str) -> dict[str, Any]:
    unit = clean(row.get("Unit")).lower()
    dimension_unit, _, weight_unit = unit.partition("/")
    original_dimensions = {
        "length": number(row.get("Length")),
        "width": number(row.get("Width")),
        "height": number(row.get("Height")),
        "unit": dimension_unit or "unknown",
    }
    primary_value = number(row.get("Weight (lb/kg)"))
    secondary_value = number(row.get("Weight (oz/g)"))
    primary_label = "lb" if weight_unit == "lb" else "kg" if weight_unit == "kg" else weight_unit
    secondary_label = "oz" if weight_unit == "oz" else "g" if weight_unit == "g" else ""
    weight_source_value = primary_value if primary_value else secondary_value
    weight_source_unit = primary_label if primary_value else secondary_label
    dims_in = {
        "length": round(convert_dimension(original_dimensions["length"], dimension_unit), 4),
        "width": round(convert_dimension(original_dimensions["width"], dimension_unit), 4),
        "height": round(convert_dimension(original_dimensions["height"], dimension_unit), 4),
        "unit": "in",
    }
    weight_lb = round(convert_weight(weight_source_value, weight_source_unit), 4)
    original = {key: serialise(value) for key, value in row.items()}
    name = clean(row.get("Name")) or clean(row.get("SKU"))
    alt_sku = clean(row.get("S-SKU"))
    source_note = f"Source unit: {unit or 'not supplied'}; original metric/European values retained in original_row."
    if alt_sku:
        source_note += f" Alternate S-SKU: {alt_sku}."
    return {
        "sku": clean(row.get("SKU")),
        "alternate_sku": alt_sku,
        "name": name,
        "description": clean(row.get("Description")),
        "client": clean(row.get("Client")),
        "dimensions_imperial": dims_in,
        "weight_lb": weight_lb,
        "original_dimensions": original_dimensions,
        "original_weight": {
            "primary_value": primary_value,
            "primary_column_unit": primary_label,
            "secondary_value": secondary_value,
            "secondary_column_unit": secondary_label,
            "source_unit": weight_unit or "unknown",
        },
        "original_unit": unit,
        "source": source,
        "source_row": source_row,
        "source_note": source_note,
        "original_row": original,
    }


def read_sku_file(source: Path, source_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    # Some exports incorrectly declare the used range as A1 even though the
    # worksheet XML contains the data rows. Force openpyxl to scan the sheet.
    sheet.reset_dimensions()
    raw_headers = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    records = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {raw_headers[index]: values[index] if index < len(values) else "" for index in range(len(raw_headers))}
        records.append(normalise_sku(row, source_row, source_name))
    units = Counter(item["original_unit"] for item in records)
    duplicates = [sku for sku, count in Counter(item["sku"] for item in records).items() if count > 1]
    return records, {
        "source_file": source.name,
        "row_count": len(records),
        "columns": raw_headers,
        "unit_distribution": dict(units),
        "duplicate_skus": duplicates,
    }


def read_skus() -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    email_records, email_meta = read_sku_file(SKU_SOURCE, "email_sku_export")
    delta_records, delta_meta = ([], {"source_file": "", "row_count": 0, "columns": [], "unit_distribution": {}, "duplicate_skus": []})
    if SKU_DELTA_SOURCE.exists():
        delta_records, delta_meta = read_sku_file(SKU_DELTA_SOURCE, "email_sku_delta_export")
    seen = set()
    all_unique = []
    overlap = []
    for record in email_records + delta_records:
        if record["sku"] in seen:
            overlap.append(record["sku"])
            continue
        seen.add(record["sku"])
        all_unique.append(record)
    return {"email_export": email_records, "email_delta": delta_records, "all_unique": all_unique}, {
        "email_export": email_meta,
        "email_delta": delta_meta,
        "duplicate_skus_across_exports": sorted(set(overlap)),
        "new_delta_skus": [record["sku"] for record in delta_records if record["sku"] not in {item["sku"] for item in email_records}],
        "unit_distribution": dict(Counter(item["original_unit"] for item in email_records + delta_records)),
    }


def read_scan_sheet() -> dict[str, Any]:
    workbook = load_workbook(SCAN_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = [clean(value) for value in next(sheet.iter_rows(values_only=True))]
    index = {header: position for position, header in enumerate(raw_headers) if header}
    rows = []
    current_inbound_date = ""
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        def get(name: str) -> Any:
            position = index.get(name)
            return values[position] if position is not None and position < len(values) else ""

        raw_date = get("Inbound Date")
        if raw_date not in (None, ""):
            current_inbound_date = serialise(raw_date)
        row = {
            "source_row": source_row,
            "inbound_date": current_inbound_date,
            "inbound_po": clean(get("Inbound PO#")),
            "outbound_date": serialise(get("Outbound date")),
            "outbound_po": clean(get("Outbound PO#")),
            "group": clean(get("Group#")),
            "sku": clean(get("SKU#")),
            "sn": clean(get("SN#")),
            "double_scan_sn": clean(get("Double-Scan SN#")),
            "location": clean(get("Location")),
            "other": clean(get("OTHER")),
            "shipout_ref": clean(get("SHIPOUT#")),
        }
        if any(row.values()):
            rows.append(row)

    inbound_rows = [row for row in rows if row["inbound_po"] and row["sku"] and row["sn"]]
    outbound_rows = [row for row in rows if row["outbound_po"] or row["outbound_date"]]
    inbound_summary: dict[str, dict[str, Any]] = {}
    for row in inbound_rows:
        po = row["inbound_po"]
        summary = inbound_summary.setdefault(po, {
            "source_inbound_po": po,
            "inbound_date": row["inbound_date"],
            "sku_quantities": Counter(),
            "serial_numbers": [],
            "source_locations": set(),
            "shipout_refs": set(),
            "source_rows": [],
        })
        summary["sku_quantities"][row["sku"]] += 1
        summary["serial_numbers"].append(row["sn"])
        if row["location"]:
            summary["source_locations"].add(row["location"])
        if row["shipout_ref"]:
            summary["shipout_refs"].add(row["shipout_ref"])
        summary["source_rows"].append(row["source_row"])
    for summary in inbound_summary.values():
        summary["sku_quantities"] = dict(summary["sku_quantities"])
        summary["source_locations"] = sorted(summary["source_locations"])
        summary["shipout_refs"] = sorted(summary["shipout_refs"])
    inbound_sn_counts = Counter(row["sn"] for row in inbound_rows)
    all_sn_counts = Counter(row["sn"] for row in rows if row["sn"])
    return {
        "source_file": SCAN_SOURCE.name,
        "source_columns": raw_headers,
        "row_count": len(rows),
        "inbound_row_count": len(inbound_rows),
        "outbound_row_count": len(outbound_rows),
        "unique_inbound_po_count": len(inbound_summary),
        "unique_inbound_sn_count": len(inbound_sn_counts),
        "unique_all_sn_count": len(all_sn_counts),
        "duplicate_inbound_sns": {sn: count for sn, count in inbound_sn_counts.items() if count > 1},
        "scan_skus": sorted({row["sku"] for row in inbound_rows}),
        "inbound_summary": sorted(inbound_summary.values(), key=lambda item: item["source_inbound_po"]),
        "inbound_rows": inbound_rows,
        "outbound_rows": outbound_rows,
    }


def build_layout() -> dict[str, Any]:
    locations = [
        {"name": "STAGE-LEFT", "zone": "Temporary Staging - Left", "type": "staging"},
        {"name": "STAGE-RIGHT", "zone": "Temporary Staging - Right", "type": "staging"},
    ]
    for zone in "ABCDE":
        for number_ in range(1, 6):
            locations.append({"name": f"{zone}{number_}", "zone": zone, "type": "floor_storage"})
    return {
        "warehouse_name": "Peak Demo Warehouse",
        "country": "United States",
        "unit_system": "US customary / imperial",
        "locations": locations,
        "location_count": len(locations),
        "layout_notes": [
            "Session requirement: two temporary staging areas, left and right.",
            "Session requirement: storage areas A1-A5, B1-B5, C1-C5, D1-D5, E1-E5.",
            "Original drawing also describes A as 7x4 and B/C as 4x4 floor grids; the 27 named locations are the current application master-data scope.",
            "Racking is near the office, four levels, one North American standard pallet per location; rack locations are not invented in this restore because their final IDs were not provided.",
        ],
    }


def supplemental_skus() -> list[dict[str, Any]]:
    def item(code: str, name: str, dims: tuple[float, float, float], weight: float, source: str, note: str) -> dict[str, Any]:
        return {
            "sku": code,
            "alternate_sku": "",
            "name": name,
            "description": note,
            "client": "Session supplemental input",
            "dimensions_imperial": {"length": dims[0], "width": dims[1], "height": dims[2], "unit": "in"},
            "weight_lb": weight,
            "original_dimensions": {"length": dims[0], "width": dims[1], "height": dims[2], "unit": "in"},
            "original_weight": {"primary_value": weight, "primary_column_unit": "lb", "secondary_value": 0, "secondary_column_unit": "", "source_unit": "lb"},
            "original_unit": "in/lb",
            "source": source,
            "source_note": note,
        }

    return [
        item("CLSCU-AA401-001", "CLSCU-AA401-001", (104, 55, 98), 3719, "session_image_3", "Screenshot record: 104 x 55 x 98 in; 3719 lb."),
        item("CLSCU-AA401-S", "CLSCU-AA401-S", (104, 55, 98), 3719, "scan_sheet_alias", "Scan-sheet SKU not present in email export; dimensions mapped from the 401 screenshot and require operational confirmation."),
        item("CLSCAC144AD702", "CLSCAC144AD702", (68, 58, 100), 3319, "session_image_4", "Screenshot record: 68 x 58 x 100 in; 3319 lb."),
        item("702-S", "702-S", (68, 58, 100), 3319, "session_request_alias", "Session-requested 702-S alias; dimensions mapped from the 702 screenshot."),
        item("401-S", "401-S", (104, 55, 98), 3719, "session_request_alias", "Session-requested 401-S alias; dimensions mapped from the 401 screenshot."),
    ]


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows({key: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value for key, value in row.items()} for row in rows)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    sku_groups, sku_meta = read_skus()
    scan = read_scan_sheet()
    layout = build_layout()
    supplemental = supplemental_skus()
    seen_skus = set()
    all_skus = []
    duplicate_import_skus = []
    for record in sku_groups["all_unique"] + supplemental:
        if record["sku"] in seen_skus:
            duplicate_import_skus.append(record["sku"])
            continue
        seen_skus.add(record["sku"])
        all_skus.append(record)

    for source in (SKU_SOURCE, SKU_SOURCE_EMPTY, SCAN_SOURCE):
        shutil.copy2(source, OUT / source.name)
    if SKU_DELTA_SOURCE.exists():
        shutil.copy2(SKU_DELTA_SOURCE, OUT / "Export Standalone_Items (1) - latest Delta SKU.xlsx")

    (OUT / "sku-master.json").write_text(json.dumps({**sku_groups, "supplemental": supplemental, "all_records": all_skus}, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "inbound-scan-history.json").write_text(json.dumps(scan, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "warehouse-layout.json").write_text(json.dumps(layout, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(OUT / "sku-master.csv", all_skus)
    write_csv(OUT / "inbound-rows.csv", scan["inbound_rows"])

    checks = {
        "sku_source_sha256": sha256(SKU_SOURCE),
        "sku_delta_source_sha256": sha256(SKU_DELTA_SOURCE) if SKU_DELTA_SOURCE.exists() else None,
        "scan_source_sha256": sha256(SCAN_SOURCE),
        "email_sku_count": len(sku_groups["email_export"]),
        "email_delta_sku_count": len(sku_groups["email_delta"]),
        "new_delta_sku_count": len(sku_meta["new_delta_skus"]),
        "supplemental_sku_count": len(supplemental),
        "total_sku_records_for_import": len(all_skus),
        "inbound_rows": scan["inbound_row_count"],
        "inbound_pos": scan["unique_inbound_po_count"],
        "unique_inbound_sns": scan["unique_inbound_sn_count"],
        "unique_all_scanned_sns": scan["unique_all_sn_count"],
        "outbound_rows_retained_but_not_imported": scan["outbound_row_count"],
        "duplicate_email_skus": sku_meta["email_export"]["duplicate_skus"],
        "duplicate_delta_skus": sku_meta["email_delta"]["duplicate_skus"],
        "duplicate_skus_across_exports": sku_meta["duplicate_skus_across_exports"],
        "duplicate_skus_all_sources": sorted(set(duplicate_import_skus)),
        "duplicate_inbound_sns": scan["duplicate_inbound_sns"],
        "unit_distribution": sku_meta["unit_distribution"],
        "conversion_rule": "cm/2.54 and mm/25.4 for dimensions; kg*2.2046226218, g/453.59237, oz/16 for weight; original values and units retained per row.",
        "import_scope": [
            "Import warehouse and 27 named locations.",
            "Import the original email SKU export, the later Delta SKU export, and explicitly marked supplemental/alias SKUs needed by the scan history.",
            "Create one system ASN per source inbound PO and preserve the original PO in the backup mapping.",
            "Import expected and received SN records; do not create stock or putaway from source locations PS-75006/PS-90670.",
            "Retain outbound rows locally only; no outbound transactions are created in this pass.",
        ],
    }
    (OUT / "verification.json").write_text(json.dumps(checks, ensure_ascii=False, indent=2), encoding="utf-8")

    readme = f"""# GreaterWMS 会话恢复包

生成日期：2026-08-07

## 来源与范围

本恢复包不读取旧数据库，只使用本次会话确认过的本地导出文件和会话输入。

- `Export Standalone_Items.xlsx`：此前邮件/SKU 主表，{len(sku_groups["email_export"])} 条。
- `Export Standalone_Items (1) - latest Delta SKU.xlsx`：后续 Delta SKU 附件，XML 实际包含 {len(sku_groups["email_delta"])} 条；该文件的工作表维度元数据错误，已按 XML 实际数据行解析。
- `Scan sheet sample.xlsx`：扫描历史，入库有效行 233 条、56 个入库 PO、233 个唯一入库 SN；全表共保留 340 个唯一 SN；出库行仅在本地保留。
- 会话附件截图：补充 401/702 的尺寸重量。

## 单位规则

系统导入值统一按美国仓库使用的英制保存：尺寸为 `in`，重量为 `lb`。每条 SKU 同时保留原始尺寸、原始重量、原始单位和来源行。原始文件没有被改写。

## 导入边界

会导入 `Peak Demo Warehouse`、两个暂存库位、A1-E5 共 27 个命名库位、SKU 主数据、按来源 PO 汇总的 ASN，以及预期/实收 SN 记录。

扫描表中的 `PS-75006`、`PS-90670` 是来源系统/客户库位，不映射到新仓库库位；否则会把历史位置误当成当前库存位置。出库历史只保存在 `inbound-scan-history.json`，本次不创建出库单。

`CLSCU-AA401-S` 是扫描表中出现但不在邮件 SKU 主表的 SKU，尺寸按 401 截图做了补充并标记为待现场确认。`702-S`、`401-S` 是会话中明确提出的别名，均保留来源标记。

## 文件

- `sku-master.json` / `sku-master.csv`：英制标准化 SKU 和原始行。
- `inbound-scan-history.json`：入库/出库扫描行、PO 汇总和 SN 重复检查。
- `warehouse-layout.json`：仓库和 27 个库位定义。
- `verification.json`：数量、哈希、换算规则和导入边界。
- 原始 `.xlsx` 文件：用于复核，不覆盖。
"""
    (OUT / "README.md").write_text(readme, encoding="utf-8")
    print(json.dumps({"backup_dir": str(OUT), **checks}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
